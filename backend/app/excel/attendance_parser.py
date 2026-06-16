"""
Parse uploaded attendance workbooks (openpyxl read-only for large files).
"""

from __future__ import annotations

import calendar
import logging
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook

from app.excel.field_utils import normalize_value, normalize_numeric_value
from app.excel.template_generator import (
    MONTHLY_DAILY_START_COL,
    MONTHLY_DATA_START_ROW,
    OVERTIME_DATA_START_ROW,
    SIGN_DAY_COUNT,
    TEMPLATE_SHEETS,
)

logger = logging.getLogger(__name__)

TRACKED_SCALAR_FIELDS = ("anomaly_summary", "supplement_submitted", "notes", "total_overtime_hours")


@dataclass
class ParsedEmployeeRow:
    name: str
    employee_code: str = ""
    department: str = ""
    position: str = ""
    anomaly_summary: str = ""
    supplement_submitted: str = ""
    notes: str = ""
    total_overtime_hours: str = ""
    daily_status: Dict[str, str] = field(default_factory=dict)

    def as_snapshot_shape(self) -> Dict[str, object]:
        payload: Dict[str, object] = {
            "employee_name": self.name,
            "employee_code": self.employee_code,
            "department": self.department,
            "position": self.position,
            "anomaly_summary": self.anomaly_summary,
            "supplement_submitted": self.supplement_submitted,
            "notes": self.notes,
            "total_overtime_hours": self.total_overtime_hours,
            "daily_status": dict(self.daily_status),
        }
        return payload


@dataclass
class ParsedWorkbook:
    year: Optional[int]
    month: Optional[int]
    employees: List[ParsedEmployeeRow]

    def employees_by_name(self) -> Dict[str, ParsedEmployeeRow]:
        return {row.name: row for row in self.employees if row.name}


def _numeric_cell_value(value) -> str:
    return normalize_numeric_value(value)


def _cell_value(value) -> str:
    return normalize_value(value)


def _parse_title_period(title: object) -> Tuple[Optional[int], Optional[int]]:
    if not title or "年" not in str(title) or "月" not in str(title):
        return None, None
    try:
        text = str(title)
        parsed_year = int(text.split("年", 1)[0])
        parsed_month = int(text.split("年", 1)[1].split("月", 1)[0])
        return parsed_year, parsed_month
    except (TypeError, ValueError):
        return None, None


def _parse_monthly_sheet(ws, year: int, month: int) -> Tuple[Tuple[Optional[int], Optional[int]], List[ParsedEmployeeRow]]:
    days_in_month = calendar.monthrange(year, month)[1]
    employees: List[ParsedEmployeeRow] = []
    inferred_period: Tuple[Optional[int], Optional[int]] = (None, None)

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_index == 2 and row:
            inferred_period = _parse_title_period(row[0])
        if row_index < MONTHLY_DATA_START_ROW:
            continue
        if not row or not row[0]:
            break
        name = str(row[0]).strip()
        if not name:
            break

        parsed = ParsedEmployeeRow(
            name=name,
            employee_code=_cell_value(row[3]) if len(row) > 3 else "",
            department=_cell_value(row[2]) if len(row) > 2 else "",
            position=_cell_value(row[4]) if len(row) > 4 else "",
            anomaly_summary=_cell_value(row[16]) if len(row) > 16 else "",
            supplement_submitted=_cell_value(row[17]) if len(row) > 17 else "",
            notes=_cell_value(row[18]) if len(row) > 18 else "",
        )

        for day in range(1, SIGN_DAY_COUNT + 1):
            if day > days_in_month:
                continue
            col_index = MONTHLY_DAILY_START_COL - 1 + day - 1
            if col_index < len(row):
                parsed.daily_status[f"day_{day}"] = _cell_value(row[col_index])

        employees.append(parsed)

    return inferred_period, employees


def _parse_overtime_sheet(ws) -> Dict[str, str]:
    overtime_by_name: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=OVERTIME_DATA_START_ROW, values_only=True):
        if not row or not row[0]:
            break
        name = str(row[0]).strip()
        if not name or name == "合计":
            break
        total_index = 7
        if len(row) > total_index and row[total_index] is not None:
            overtime_by_name[name] = _numeric_cell_value(row[total_index])
        elif len(row) > 4 and row[4] is not None:
            overtime_by_name[name] = _numeric_cell_value(row[4])
    return overtime_by_name


def parse_attendance_workbook(
    source: Union[str, Path, BinaryIO, BytesIO],
    *,
    year: int,
    month: int,
) -> ParsedWorkbook:
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        if TEMPLATE_SHEETS[2] not in wb.sheetnames:
            raise ValueError("Uploaded file is missing the 月度汇总 worksheet")

        monthly_ws = wb[TEMPLATE_SHEETS[2]]
        (inferred_year, inferred_month), employees = _parse_monthly_sheet(monthly_ws, year, month)

        if TEMPLATE_SHEETS[3] in wb.sheetnames:
            overtime_map = _parse_overtime_sheet(wb[TEMPLATE_SHEETS[3]])
            for employee in employees:
                if employee.name in overtime_map:
                    employee.total_overtime_hours = overtime_map[employee.name]

        return ParsedWorkbook(
            year=inferred_year or year,
            month=inferred_month or month,
            employees=employees,
        )
    finally:
        wb.close()
