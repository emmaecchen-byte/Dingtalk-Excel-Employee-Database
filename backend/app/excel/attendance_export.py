"""
Populate the master attendance Excel template from database records.
"""

from __future__ import annotations

import calendar
import logging
import os
import tempfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Sequence, Tuple, Union

from openpyxl import load_workbook
from sqlalchemy.orm import Session, joinedload

from app.excel.template_generator import (
    MONTHLY_DAILY_START_COL,
    MONTHLY_DATA_START_ROW,
    MONTHLY_META_ANOMALY_COL,
    MONTHLY_META_NOTES_COL,
    MONTHLY_META_SUPPLEMENT_COL,
    OVERTIME_CALC_START_COL,
    OVERTIME_DATA_START_ROW,
    OVERTIME_DAY_START_COL,
    SIGN_DATA_START_ROW,
    SIGN_DAY_COUNT,
    SIGN_DAY_START_COL,
    SITUATION_DATA_START_ROW,
    TEMPLATE_SHEETS,
    TemplateEmployee,
    build_attendance_workbook,
    generate_attendance_template,
)
from app.models import MonthlyAttendance

logger = logging.getLogger(__name__)

ANOMALY_SYMBOLS = frozenset({"※", "●", "缺", "×", "迟"})
HALF_DAY_SEPARATORS = ("/", "|", "、")
DEFAULT_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2] / "templates" / "master_template.xlsx"
)


@dataclass
class ExcelExportResult:
    """Generated attendance workbook."""

    path: Path
    filename: str
    year: int
    month: int
    employee_count: int

    def as_stream(self) -> BytesIO:
        buffer = BytesIO(self.path.read_bytes())
        buffer.seek(0)
        return buffer

    def cleanup(self) -> None:
        try:
            self.path.unlink(missing_ok=True)
        except OSError:
            logger.warning("Failed to remove temporary Excel file: %s", self.path)


class AttendanceExcelError(Exception):
    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


def _record_to_template_employee(record: MonthlyAttendance) -> TemplateEmployee:
    employee = record.employee
    return TemplateEmployee(
        name=employee.name,
        department=employee.department or "",
        position=employee.position or "",
        employee_code=employee.employee_code or "",
    )


def get_day_value(record: MonthlyAttendance, day: int) -> str:
    field_name = f"day_{day}"
    overrides = record.manual_overrides or {}
    if field_name in overrides and overrides[field_name] is not None:
        return str(overrides[field_name]).strip()
    value = getattr(record, field_name, None)
    return str(value).strip() if value else ""


def split_day_halves(value: str) -> Tuple[str, str]:
    if not value:
        return "", ""
    text = value.strip()
    for separator in HALF_DAY_SEPARATORS:
        if separator in text:
            left, right = text.split(separator, 1)
            return left.strip(), right.strip()
    if len(text) == 2 and not any(separator in text for separator in HALF_DAY_SEPARATORS):
        return text[0], text[1]
    return text, text


def combined_day_status(am_value: str, pm_value: str) -> str:
    if am_value and pm_value and am_value == pm_value:
        return am_value
    if am_value and pm_value:
        return f"{am_value}{pm_value}"
    return am_value or pm_value


def _has_anomaly_in_day(value: str) -> bool:
    if not value:
        return False
    am_value, pm_value = split_day_halves(value)
    return am_value in ANOMALY_SYMBOLS or pm_value in ANOMALY_SYMBOLS


def first_anomaly_date(record: MonthlyAttendance, year: int, month: int) -> Optional[str]:
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        if _has_anomaly_in_day(get_day_value(record, day)):
            return f"{year}-{month:02d}-{day:02d}"
    return None


def resolve_day_value(record: MonthlyAttendance, day: int) -> str:
    value = get_day_value(record, day)
    if value:
        return value
    return ""


def _query_attendance_records(
    db: Session,
    company_id: int,
    year: int,
    month: int,
) -> List[MonthlyAttendance]:
    return (
        db.query(MonthlyAttendance)
        .options(joinedload(MonthlyAttendance.employee))
        .filter(
            MonthlyAttendance.company_id == company_id,
            MonthlyAttendance.year == year,
            MonthlyAttendance.month == month,
        )
        .order_by(MonthlyAttendance.employee_id)
        .all()
    )


def _count_template_employee_slots(sign_ws) -> int:
    """Count 上午/下午 row pairs in the sign-off sheet (including blank name slots)."""
    count = 0
    row = SIGN_DATA_START_ROW
    while True:
        if sign_ws.cell(row=row, column=3).value != "上午":
            break
        if sign_ws.cell(row=row + 1, column=3).value != "下午":
            break
        count += 1
        row += 2
    return count


def _open_workbook(
    year: int,
    month: int,
    employees: Sequence[TemplateEmployee],
    template_path: Optional[Path],
):
    path = template_path or DEFAULT_TEMPLATE_PATH
    employee_count = len(employees)
    if path.exists():
        workbook = load_workbook(path)
        if workbook.sheetnames == list(TEMPLATE_SHEETS):
            sign_ws = workbook[TEMPLATE_SHEETS[0]]
            slot_count = _count_template_employee_slots(sign_ws)
            if slot_count >= employee_count:
                logger.info(
                    "Loaded master template from %s (%s slots, %s employees)",
                    path,
                    slot_count,
                    employee_count,
                )
                return workbook
        logger.info(
            "Rebuilding workbook for %s employees (template mismatch or insufficient slots)",
            employee_count,
        )
    else:
        logger.info("Master template not found at %s; building workbook", path)

    return build_attendance_workbook(year, month, employees)


def _populate_sign_sheet(ws, records: Sequence[MonthlyAttendance], year: int, month: int) -> None:
    days_in_month = calendar.monthrange(year, month)[1]
    for index, record in enumerate(records):
        am_row = SIGN_DATA_START_ROW + index * 2
        pm_row = am_row + 1
        ws.cell(row=am_row, column=2, value=record.employee.name)

        for day in range(1, SIGN_DAY_COUNT + 1):
            col = SIGN_DAY_START_COL + day - 1
            if day > days_in_month:
                continue
            day_value = resolve_day_value(record, day)
            am_value, pm_value = split_day_halves(day_value)
            ws.cell(row=am_row, column=col, value=am_value)
            ws.cell(row=pm_row, column=col, value=pm_value)


def _populate_monthly_sheet(ws, records: Sequence[MonthlyAttendance], year: int, month: int) -> None:
    days_in_month = calendar.monthrange(year, month)[1]
    for index, record in enumerate(records):
        row = MONTHLY_DATA_START_ROW + index
        employee = record.employee
        ws.cell(row=row, column=1, value=employee.name)
        ws.cell(row=row, column=2, value="默认考勤组")
        ws.cell(row=row, column=3, value=employee.department or "")
        ws.cell(row=row, column=4, value=employee.employee_code or "")
        ws.cell(row=row, column=5, value=employee.position or "")
        ws.cell(row=row, column=MONTHLY_META_ANOMALY_COL, value=record.anomaly_summary or "")
        ws.cell(row=row, column=MONTHLY_META_SUPPLEMENT_COL, value="是" if record.supplement_submitted else "否")
        ws.cell(row=row, column=MONTHLY_META_NOTES_COL, value=record.notes or "")

        for day in range(1, SIGN_DAY_COUNT + 1):
            col = MONTHLY_DAILY_START_COL + day - 1
            if day > days_in_month:
                continue
            day_value = resolve_day_value(record, day)
            am_value, pm_value = split_day_halves(day_value)
            ws.cell(row=row, column=col, value=combined_day_status(am_value, pm_value))


def _populate_situation_sheet(ws, records: Sequence[MonthlyAttendance], year: int, month: int) -> None:
    row = SITUATION_DATA_START_ROW
    for record in records:
        if not (
            record.absenteeism_count > 0
            or record.lateness_count > 0
            or record.missing_punch_count > 0
        ):
            continue

        ws.cell(row=row, column=1, value=record.employee.name)
        ws.cell(row=row, column=2, value=first_anomaly_date(record, year, month) or f"{year}-{month:02d}-01")
        ws.cell(row=row, column=3, value=record.anomaly_summary or "")
        ws.cell(row=row, column=4, value="是" if record.supplement_submitted else "否")
        ws.cell(row=row, column=5, value=record.notes or "")
        row += 1


def _populate_overtime_sheet(ws, records: Sequence[MonthlyAttendance]) -> None:
    total_col = OVERTIME_CALC_START_COL + 3  # AM — 加班合计
    for index, record in enumerate(records):
        row = OVERTIME_DATA_START_ROW + index
        employee = record.employee
        overtime_hours = float(record.total_overtime_hours or 0)
        ws.cell(row=row, column=1, value=employee.name)
        ws.cell(row=row, column=2, value=employee.department or "")
        ws.cell(row=row, column=3, value="调休")
        if overtime_hours:
            ws.cell(row=row, column=OVERTIME_DAY_START_COL, value=overtime_hours)
        ws.cell(row=row, column=total_col, value=overtime_hours)


def populate_workbook(
    workbook,
    records: Sequence[MonthlyAttendance],
    year: int,
    month: int,
) -> None:
    sign_ws = workbook[TEMPLATE_SHEETS[0]]
    situation_ws = workbook[TEMPLATE_SHEETS[1]]
    monthly_ws = workbook[TEMPLATE_SHEETS[2]]
    overtime_ws = workbook[TEMPLATE_SHEETS[3]]

    _populate_sign_sheet(sign_ws, records, year, month)
    _populate_monthly_sheet(monthly_ws, records, year, month)
    _populate_situation_sheet(situation_ws, records, year, month)
    _populate_overtime_sheet(overtime_ws, records)


def generate_attendance_excel(
    db: Session,
    company_id: int,
    year: int,
    month: int,
    *,
    template_path: Optional[Union[str, Path]] = None,
    output_dir: Optional[Union[str, Path]] = None,
) -> ExcelExportResult:
    """
    Load the master template, populate it from monthly_attendance, and save to a temp file.

    Returns an ExcelExportResult with the file path and helpers for streaming download.
    """
    records = _query_attendance_records(db, company_id, year, month)
    if not records:
        raise AttendanceExcelError(f"No attendance data for {year}-{month:02d}")

    employees = [_record_to_template_employee(record) for record in records]
    resolved_template = Path(template_path) if template_path else None

    workbook = _open_workbook(year, month, employees, resolved_template)
    populate_workbook(workbook, records, year, month)

    filename = f"attendance_{year}_{month:02d}.xlsx"
    if output_dir:
        target_dir = Path(output_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        output_path = target_dir / filename
        workbook.save(output_path)
    else:
        fd, temp_name = tempfile.mkstemp(prefix="attendance_", suffix=".xlsx")
        os.close(fd)
        output_path = Path(temp_name)
        workbook.save(output_path)

    logger.info(
        "Generated attendance Excel for company_id=%s %s-%02d (%s employees) -> %s",
        company_id,
        year,
        month,
        len(records),
        output_path,
    )

    return ExcelExportResult(
        path=output_path,
        filename=filename,
        year=year,
        month=month,
        employee_count=len(records),
    )


def ensure_master_template(path: Optional[Union[str, Path]] = None) -> Path:
    """Create the master template on disk if it does not exist."""
    target = Path(path) if path else DEFAULT_TEMPLATE_PATH
    if not target.exists():
        generate_attendance_template(target, year=datetime.utcnow().year, month=datetime.utcnow().month)
    return target
