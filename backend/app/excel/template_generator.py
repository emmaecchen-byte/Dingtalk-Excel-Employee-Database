"""
Master attendance Excel template generator (openpyxl).

Sheets:
  1. 签字           — Daily sign-off grid with COUNTIF summary formulas (AJ–AR, AT)
  2. 情况说明       — Anomaly explanations (populated by backend, no formulas)
  3. 月度汇总       — Monthly summary: name in A, daily columns B–AF
  4. 加班结算加班工资 — Daily overtime D–AH and rate columns AJ–AP
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Sequence, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COMPANY_NAME = "创崎新能源技术（上海）有限公司"

TEMPLATE_SHEETS = ("签字", "情况说明", "月度汇总", "加班结算加班工资")

# 签字 sheet layout
SIGN_HEADER_ROW = 3
SIGN_DATE_ROW = 4
SIGN_LEGEND_ROW = 5
SIGN_DATA_START_ROW = 6

SIGN_DAY_START_COL = 4  # D
SIGN_DAY_END_COL = 34  # AH (31 days)
SIGN_DAY_COUNT = 31

SIGN_SUMMARY_START_COL = 36  # AJ
SIGN_SUMMARY_END_COL = 44  # AR
SIGN_LEGEND_END_COL = 45  # AS (AJ5–AS5)
SIGN_ABSENT_COL = 46  # AT = 应出勤 - 出勤

# Legend symbols in AJ5–AS5 (symbol only in cell; label in comment / adjacent doc)
SIGN_LEGEND_SYMBOLS: Tuple[Tuple[str, str], ...] = (
    ("√", "出勤"),
    ("◇", "事假"),
    ("✬", "调休"),
    ("▼", "出差"),
    ("※", "病假"),
    ("●", "福利假"),
    ("AL", "年假"),
    ("○", "产假/陪产假"),
    ("FL", "丧假"),
    ("ML", "婚假"),
)

SIGN_COUNT_SYMBOLS: Tuple[str, ...] = tuple(symbol for symbol, _ in SIGN_LEGEND_SYMBOLS[:9])

# 月度汇总 — name in A, daily status B–AF (31 days)
MONTHLY_TITLE_ROW = 1
MONTHLY_GENERATED_ROW = 2
MONTHLY_SECTION_ROW = 3
MONTHLY_HEADER_ROW = 4
MONTHLY_DATA_START_ROW = 5
MONTHLY_INFO_END_COL = 1  # A = 姓名
MONTHLY_DAILY_START_COL = 2  # B = day 1
MONTHLY_DAILY_END_COL = 32  # AF = day 31
MONTHLY_DATE_HEADER_ROW = 1  # B1–AF1 day-of-month / 六 / 日
# Legacy parser columns (not written to web-export workbooks)
MONTHLY_META_ANOMALY_COL = 33  # AG
MONTHLY_META_SUPPLEMENT_COL = 34  # AH
MONTHLY_META_NOTES_COL = 35  # AI

# 情况说明
SITUATION_DATA_START_ROW = 2

# 加班结算
OVERTIME_TITLE_ROW = 1
OVERTIME_MAIN_HEADER_ROW = 2
OVERTIME_RATE_HEADER_ROW = 4
OVERTIME_HEADER_ROW = OVERTIME_RATE_HEADER_ROW  # backward-compatible alias
OVERTIME_DATA_START_ROW = 5
OVERTIME_DAY_START_COL = 4  # D
OVERTIME_DAY_COUNT = 31
OVERTIME_CALC_START_COL = 36  # AJ
OVERTIME_CALC_END_COL = 42  # AP

# 调休 section (compensatory time-off) — mirrors original rows 90–186
OVERTIME_COMP_HEADER_ROW = 90
OVERTIME_COMP_SUBHEADER_ROW = 92
OVERTIME_COMP_DATA_START_ROW = 93
OVERTIME_COMP_MAX_ROW = 186
OVERTIME_COMP_PRIOR_COL = 4  # D = 过往加班
OVERTIME_COMP_DAY_START_COL = 5  # E = day 1 (D is prior-month balance column)
OVERTIME_COMP_DAY_END_COL = 35  # AI = day 31
OVERTIME_COMP_PREV_REMAIN_COL = 36  # AJ = prior month remaining hours
OVERTIME_COMP_MONTH_HOURS_COL = 37  # AK = 1x month total (=SUM(D:AI))
OVERTIME_COMP_REMAIN_COL = 42  # AP = AJ + AK

OVERTIME_SETTLEMENT_PAY = "加班费"
OVERTIME_SETTLEMENT_COMP = "调休"

# 1.5x / 2x / 3x overtime day column groups (within D–AH)
OVERTIME_15X_RANGES = ("J:M", "O:S", "V:Z", "AC:AG")
OVERTIME_2X_RANGES = ("G:I", "N:N", "T:U", "AA:AB", "AH:AI")
OVERTIME_3X_RANGES = ("E:F",)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EMPTY_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
OVERTIME_SUMMARY_DATA_FILL = PatternFill("solid", fgColor="90EE90")
OVERTIME_FORMULA_FONT = Font(name="宋体", size=10, bold=True, color="FF0000")
WEEKEND_FILL = PatternFill("solid", fgColor="D9D9D9")
OUT_OF_MONTH_FILL = PatternFill("solid", fgColor="F2F2F2")
TITLE_FONT = Font(name="宋体", size=14, bold=True)
HEADER_FONT = Font(name="宋体", size=10, bold=True)
BODY_FONT = Font(name="宋体", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

MONTHLY_INFO_COL_WIDTH = 15
MONTHLY_DAILY_COL_WIDTH = 10

DEFAULT_BLANK_EMPLOYEE_SLOTS = 15


@dataclass
class TemplateEmployee:
    name: str
    department: str = ""
    position: str = ""
    employee_code: str = ""
    attendance_group: str = "默认考勤组"
    overtime_settlement: str = "调休"


def _col_letter(col: int) -> str:
    return get_column_letter(col)


def _day_col_sign(day: int) -> str:
    return _col_letter(SIGN_DAY_START_COL + day - 1)


def _summary_col_sign(index: int) -> str:
    return _col_letter(SIGN_SUMMARY_START_COL + index)


def _monthly_daily_col(day: int) -> str:
    return _col_letter(MONTHLY_DAILY_START_COL + day - 1)


def monthly_day_header_label(year: int, month: int, day: int) -> str:
    """Weekday number, 六 for Saturday, 日 for Sunday."""
    weekday = date(year, month, day).weekday()
    if weekday == 5:
        return "六"
    if weekday == 6:
        return "日"
    return str(day)


def is_calendar_weekend(year: int, month: int, day: int) -> bool:
    return date(year, month, day).weekday() >= 5


def count_month_work_days(year: int, month: int) -> int:
    """Count Monday–Friday days in a calendar month (应出勤天数)."""
    days_in_month = calendar.monthrange(year, month)[1]
    return sum(
        1
        for day in range(1, days_in_month + 1)
        if not is_calendar_weekend(year, month, day)
    )


def monthly_stats_title(year: int, month: int) -> str:
    days_in_month = calendar.monthrange(year, month)[1]
    return (
        f"月度汇总 统计日期：{year}-{month:02d}-01 至 {year}-{month:02d}-{days_in_month:02d}"
    )


def write_monthly_summary_title_row(ws, year: int, month: int) -> None:
    """Write row 1 title in A1 for the 月度汇总 sheet."""
    title = monthly_stats_title(year, month)
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row <= MONTHLY_TITLE_ROW <= merged.max_row
            and merged.min_col <= 1 <= merged.max_col
        ):
            ws.unmerge_cells(str(merged))

    ws.merge_cells(
        start_row=MONTHLY_TITLE_ROW,
        start_column=1,
        end_row=MONTHLY_TITLE_ROW,
        end_column=MONTHLY_INFO_END_COL,
    )
    cell = ws.cell(row=MONTHLY_TITLE_ROW, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[MONTHLY_TITLE_ROW].height = 24


def monthly_generated_label(generated_at: Optional[datetime] = None) -> str:
    stamp = generated_at or datetime.utcnow()
    return f"报表生成时间：{stamp.strftime('%Y-%m-%d %H:%M')}"


def write_monthly_summary_timestamp_row(
    ws,
    generated_at: Optional[datetime] = None,
) -> None:
    """Write row 2 timestamp in A2 for the 月度汇总 sheet."""
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row <= MONTHLY_GENERATED_ROW <= merged.max_row
            and merged.min_col <= 1 <= merged.max_col
        ):
            ws.unmerge_cells(str(merged))

    ws.merge_cells(
        start_row=MONTHLY_GENERATED_ROW,
        start_column=1,
        end_row=MONTHLY_GENERATED_ROW,
        end_column=MONTHLY_INFO_END_COL,
    )
    cell = ws.cell(
        row=MONTHLY_GENERATED_ROW,
        column=1,
        value=monthly_generated_label(generated_at),
    )
    cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_monthly_day_column_headers(
    ws,
    year: int,
    month: int,
    header_row: int,
) -> None:
    """
    Write B–AF date headers for a month.

    Weekdays show the day number (1–31). Saturday → 六, Sunday → 日.
    """
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, SIGN_DAY_COUNT + 1):
        col = MONTHLY_DAILY_START_COL + day - 1
        if day <= days_in_month:
            header_value = monthly_day_header_label(year, month, day)
            is_weekend = is_calendar_weekend(year, month, day)
        else:
            header_value = f"{day}*"
            is_weekend = False
        cell = ws.cell(row=header_row, column=col, value=header_value)
        cell.font = HEADER_FONT
        cell.fill = WEEKEND_FILL if is_weekend else HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if day > days_in_month:
            cell.font = Font(name="宋体", size=10, bold=True, color="999999")
            cell.fill = OUT_OF_MONTH_FILL


def configure_monthly_summary_headers(
    ws,
    year: int,
    month: int,
    *,
    generated_at: Optional[datetime] = None,
) -> None:
    """Apply rows 1–4 for the 月度汇总 sheet (title, timestamp, section labels, column headers)."""
    days_in_month = calendar.monthrange(year, month)[1]

    write_monthly_summary_title_row(ws, year, month)
    write_monthly_summary_timestamp_row(ws, generated_at=generated_at)

    # Row 1: B1–AF1 date headers alongside title in A1
    write_monthly_day_column_headers(ws, year, month, MONTHLY_DATE_HEADER_ROW)

    ws.cell(row=MONTHLY_SECTION_ROW, column=1, value="姓名")
    section = ws.cell(row=MONTHLY_SECTION_ROW, column=1)
    section.font = HEADER_FONT
    section.fill = HEADER_FILL
    section.alignment = CENTER

    ws.merge_cells(
        start_row=MONTHLY_SECTION_ROW,
        start_column=MONTHLY_DAILY_START_COL,
        end_row=MONTHLY_SECTION_ROW,
        end_column=MONTHLY_DAILY_END_COL,
    )
    daily_section = ws.cell(row=MONTHLY_SECTION_ROW, column=MONTHLY_DAILY_START_COL, value="每日考勤状态")
    daily_section.font = HEADER_FONT
    daily_section.fill = HEADER_FILL
    daily_section.alignment = CENTER

    ws.cell(row=MONTHLY_HEADER_ROW, column=1, value="姓名")
    name_header = ws.cell(row=MONTHLY_HEADER_ROW, column=1)
    name_header.font = HEADER_FONT
    name_header.fill = HEADER_FILL
    name_header.alignment = LEFT
    name_header.border = THIN_BORDER

    # Row 4: repeat date headers under column header row for the data table
    write_monthly_day_column_headers(ws, year, month, MONTHLY_HEADER_ROW)


def apply_monthly_column_widths(ws) -> None:
    """Set 月度汇总 column widths: A=15, B–AF=10."""
    for col in range(1, MONTHLY_INFO_END_COL + 1):
        ws.column_dimensions[_col_letter(col)].width = MONTHLY_INFO_COL_WIDTH
    for col in range(MONTHLY_DAILY_START_COL, MONTHLY_DAILY_END_COL + 1):
        ws.column_dimensions[_col_letter(col)].width = MONTHLY_DAILY_COL_WIDTH


def prepare_monthly_spacer_columns(ws, *, last_row: int) -> None:
    """No-op: legacy spacer columns between employee info and daily status were removed."""
    del ws, last_row


def format_monthly_summary_sheet(
    ws,
    year: int,
    month: int,
    *,
    last_row: int,
) -> None:
    """
    Apply borders, alignment, and weekend column highlighting to 月度汇总.

    Name column (A): left-aligned. Daily headers and status (B–AF): centered.
    Weekend columns (六/日): gray background.
    """
    days_in_month = calendar.monthrange(year, month)[1]
    apply_monthly_column_widths(ws)

    for row in range(MONTHLY_TITLE_ROW, last_row + 1):
        for col in range(1, MONTHLY_DAILY_END_COL + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER

            if col <= MONTHLY_INFO_END_COL:
                if row == MONTHLY_SECTION_ROW:
                    cell.alignment = CENTER
                elif row in (MONTHLY_TITLE_ROW, MONTHLY_GENERATED_ROW):
                    cell.alignment = LEFT
                    if row == MONTHLY_TITLE_ROW and col == 1:
                        cell.font = TITLE_FONT
                    elif row == MONTHLY_GENERATED_ROW and col == 1:
                        cell.font = BODY_FONT
                else:
                    cell.alignment = LEFT
                    if row == MONTHLY_HEADER_ROW:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                    elif row >= MONTHLY_DATA_START_ROW:
                        cell.font = BODY_FONT
                continue

            day = col - MONTHLY_DAILY_START_COL + 1
            cell.alignment = CENTER
            has_status_value = row >= MONTHLY_DATA_START_ROW and bool(cell.value)
            if has_status_value:
                continue
            if day > days_in_month:
                cell.fill = OUT_OF_MONTH_FILL
            elif is_calendar_weekend(year, month, day):
                cell.fill = WEEKEND_FILL
                if row in (MONTHLY_DATE_HEADER_ROW, MONTHLY_HEADER_ROW):
                    cell.font = HEADER_FONT
            elif row in (MONTHLY_DATE_HEADER_ROW, MONTHLY_HEADER_ROW):
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    prepare_monthly_spacer_columns(ws, last_row=last_row)


def _overtime_day_col(day: int) -> str:
    return _col_letter(OVERTIME_DAY_START_COL + day - 1)


def _apply_border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _merge_title(ws, row: int, min_col: int, max_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=min_col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = CENTER


def sign_countif_formula(data_row: int, legend_col_index: int) -> str:
    """COUNTIF for one summary column (AJ=0 … AR=8), referencing row 5 legend."""
    day_start = _day_col_sign(1)
    day_end = _day_col_sign(SIGN_DAY_COUNT)
    legend_col = _summary_col_sign(legend_col_index)
    return f"=COUNTIF({day_start}{data_row}:{day_end}{data_row},${legend_col}$5)"


def sign_absent_formula(data_row: int, work_days: int) -> str:
    """AT column: working days minus attendance count in AJ (缺勤 = 应出勤 - 出勤)."""
    aj_col = _summary_col_sign(0)
    return f"={work_days}-{aj_col}{data_row}"


def write_sign_sheet_employee_summary_formulas(
    ws,
    data_row: int,
    *,
    work_days: int,
) -> None:
    """Write AJ–AR COUNTIF formulas and AT absent formula for one 签字 data row."""
    for idx in range(len(SIGN_COUNT_SYMBOLS)):
        col = SIGN_SUMMARY_START_COL + idx
        cell = ws.cell(row=data_row, column=col, value=sign_countif_formula(data_row, idx))
        cell.alignment = CENTER

    absent_cell = ws.cell(
        row=data_row,
        column=SIGN_ABSENT_COL,
        value=sign_absent_formula(data_row, work_days),
    )
    absent_cell.alignment = CENTER


def write_sign_sheet_employee_am_pm_summary_formulas(
    ws,
    am_row: int,
    pm_row: int,
    *,
    work_days: int,
) -> None:
    """Write COUNTIF and absent formulas on both 上午 and 下午 rows for one employee."""
    for row in (am_row, pm_row):
        write_sign_sheet_employee_summary_formulas(ws, row, work_days=work_days)


def write_sign_sheet_summary_formulas(
    ws,
    employee_count: int,
    year: int,
    month: int,
) -> None:
    """Write COUNTIF and AT formulas for every employee on the 签字 sheet (rows 6–7, 8–9, …)."""
    work_days = count_month_work_days(year, month)
    for index in range(employee_count):
        am_row = SIGN_DATA_START_ROW + index * 2
        pm_row = am_row + 1
        write_sign_sheet_employee_am_pm_summary_formulas(ws, am_row, pm_row, work_days=work_days)


def _sign_countif_formula(am_row: int, symbol: str) -> str:
    return sign_countif_formula(am_row, SIGN_COUNT_SYMBOLS.index(symbol))


def _overtime_sum_formula(row: int, ranges: Tuple[str, ...]) -> str:
    parts = []
    for rng in ranges:
        if ":" not in rng:
            parts.append(f"{rng}{row}")
            continue
        start, end = rng.split(":", 1)
        if start == end:
            parts.append(f"{start}{row}")
        else:
            parts.append(f"{start}{row}:{end}{row}")
    return f"=SUM({','.join(parts)})"


def overtime_15x_hours_formula(row: int) -> str:
    return _overtime_sum_formula(row, OVERTIME_15X_RANGES)


def overtime_2x_hours_formula(row: int) -> str:
    return _overtime_sum_formula(row, OVERTIME_2X_RANGES)


def overtime_3x_hours_formula(row: int) -> str:
    return _overtime_sum_formula(row, OVERTIME_3X_RANGES)


def overtime_15x_total_formula(row: int) -> str:
    return f"=AJ{row}*1.5"


def overtime_2x_total_formula(row: int) -> str:
    return f"=AK{row}*2"


def overtime_3x_total_formula(row: int) -> str:
    return f"=AL{row}*3"


def overtime_pay_total_formula(row: int) -> str:
    return f"=SUM(AM{row}:AO{row})"


def _prev_month_year_month(year: int, month: int) -> Tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


def overtime_pay_summary_labels(month: int) -> dict:
    return {
        "hours_total": f"{month}月加班时长合计",
        "multiplier_total": f"{month}月加班倍数合计",
        "grand_total": "总计",
    }


def _style_overtime_summary_header_cell(ws, row: int, col: int, value) -> None:
    """Summary title text — no green fill (matches original header boxes)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = NO_FILL
    cell.alignment = CENTER


def _clear_summary_header_merge_fill(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """Remove inherited green/data fills from merged summary header cells."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).fill = NO_FILL


def _style_overtime_summary_header_merge(
    ws,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    value: str,
) -> None:
    _clear_summary_header_merge_fill(ws, min_row, max_row, min_col, max_col)
    _style_overtime_summary_header_cell(ws, min_row, min_col, value)


def _style_overtime_summary_data_cell(cell) -> None:
    """Light-green background with bold red values in AJ–AP data columns."""
    cell.fill = OVERTIME_SUMMARY_DATA_FILL
    if cell.value not in (None, ""):
        cell.font = OVERTIME_FORMULA_FONT


def _apply_overtime_summary_column_style(ws, row: int) -> None:
    for col in range(OVERTIME_CALC_START_COL, OVERTIME_CALC_END_COL + 1):
        _style_overtime_summary_data_cell(ws.cell(row=row, column=col))


def _style_overtime_formula_cell(cell) -> None:
    _style_overtime_summary_data_cell(cell)


def write_overtime_pay_section_headers(ws, year: int, month: int) -> None:
    """
    Write the 加班费 header block (rows 2–4) on 加班结算加班工资.

    Matches the original workbook: title headers on row 2 and 1.5×/2×/3× labels on row 4.
    Data cells in AJ–AP use light-green fill with bold red numbers.
    """
    labels = overtime_pay_summary_labels(month)
    main_row = OVERTIME_MAIN_HEADER_ROW
    sub_row = 3
    rate_row = OVERTIME_RATE_HEADER_ROW
    days_in_month = calendar.monthrange(year, month)[1]

    _unmerge_rows(ws, main_row, rate_row)

    for col, label in (
        (1, "姓名"),
        (2, "部门"),
        (3, "加班兑换方式"),
        (4, "过往\n加班"),
    ):
        ws.merge_cells(
            start_row=main_row,
            start_column=col,
            end_row=rate_row,
            end_column=col,
        )
        _style_overtime_header_cell(ws, main_row, col, label)

    for day in range(1, OVERTIME_DAY_COUNT + 1):
        col = OVERTIME_DAY_START_COL + day - 1
        ws.merge_cells(
            start_row=main_row,
            start_column=col,
            end_row=rate_row,
            end_column=col,
        )
        value = day if day <= days_in_month else f"{day}*"
        _style_overtime_header_cell(ws, main_row, col, value)
        if day > days_in_month:
            ws.cell(row=main_row, column=col).font = Font(
                name="宋体", size=10, bold=True, color="999999"
            )

    ws.merge_cells(
        start_row=main_row,
        start_column=OVERTIME_CALC_START_COL,
        end_row=sub_row,
        end_column=OVERTIME_CALC_START_COL + 2,
    )
    _style_overtime_summary_header_merge(
        ws,
        main_row,
        OVERTIME_CALC_START_COL,
        sub_row,
        OVERTIME_CALC_START_COL + 2,
        labels["hours_total"],
    )

    ws.merge_cells(
        start_row=main_row,
        start_column=OVERTIME_CALC_START_COL + 3,
        end_row=sub_row,
        end_column=OVERTIME_CALC_START_COL + 5,
    )
    _style_overtime_summary_header_merge(
        ws,
        main_row,
        OVERTIME_CALC_START_COL + 3,
        sub_row,
        OVERTIME_CALC_START_COL + 5,
        labels["multiplier_total"],
    )

    ws.merge_cells(
        start_row=main_row,
        start_column=OVERTIME_CALC_END_COL,
        end_row=rate_row,
        end_column=OVERTIME_CALC_END_COL,
    )
    _style_overtime_summary_header_merge(
        ws,
        main_row,
        OVERTIME_CALC_END_COL,
        rate_row,
        OVERTIME_CALC_END_COL,
        labels["grand_total"],
    )

    for idx, label in enumerate(("1.5倍", "2倍", "3倍", "1.5倍", "2倍", "3倍")):
        col = OVERTIME_CALC_START_COL + idx
        cell = ws.cell(row=rate_row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def overtime_comp_header_labels(year: int, month: int) -> dict:
    _, prev_month = _prev_month_year_month(year, month)
    return {
        "month_total": f"{month}月加班时长合计",
        "month_deduct": f"{month}月加班抵扣合计",
        "month_remain": f"{month}月剩余",
        "prev_remain": f"{prev_month}月剩余",
    }


def overtime_comp_month_hours_formula(row: int) -> str:
    """AK: sum of 过往加班 (D) and daily hours (E–AI)."""
    return f"=SUM(D{row}:AI{row})"


def overtime_comp_remain_formula(row: int) -> str:
    """AP: prior-month balance (AJ) plus current-month total (AK)."""
    return f"=AJ{row}+AK{row}"


def _style_overtime_header_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER


def _unmerge_rows(ws, min_row: int, max_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= max_row and merged.max_row >= min_row:
            ws.unmerge_cells(str(merged))


def write_overtime_compensatory_headers(ws, year: int, month: int) -> None:
    """Write the three-row 调休 header block (rows 90–92) on 加班结算加班工资."""
    labels = overtime_comp_header_labels(year, month)
    header_row = OVERTIME_COMP_HEADER_ROW
    sub_row = OVERTIME_COMP_SUBHEADER_ROW
    days_in_month = calendar.monthrange(year, month)[1]

    _unmerge_rows(ws, header_row, sub_row)

    for col, label in (
        (1, "姓名"),
        (2, "部门"),
        (3, "加班兑换方式"),
        (4, "过往\n加班"),
    ):
        ws.merge_cells(
            start_row=header_row,
            start_column=col,
            end_row=sub_row,
            end_column=col,
        )
        _style_overtime_header_cell(ws, header_row, col, label)

    for day in range(1, OVERTIME_DAY_COUNT + 1):
        col = OVERTIME_COMP_DAY_START_COL + day - 1
        ws.merge_cells(
            start_row=header_row,
            start_column=col,
            end_row=sub_row,
            end_column=col,
        )
        value = day if day <= days_in_month else f"{day}*"
        _style_overtime_header_cell(ws, header_row, col, value)
        if day > days_in_month:
            ws.cell(row=header_row, column=col).font = Font(
                name="宋体", size=10, bold=True, color="999999"
            )

    ws.merge_cells(
        start_row=header_row,
        start_column=OVERTIME_COMP_PREV_REMAIN_COL,
        end_row=header_row + 1,
        end_column=OVERTIME_COMP_PREV_REMAIN_COL + 2,
    )
    _style_overtime_summary_header_merge(
        ws,
        header_row,
        OVERTIME_COMP_PREV_REMAIN_COL,
        header_row + 1,
        OVERTIME_COMP_PREV_REMAIN_COL + 2,
        labels["month_total"],
    )

    ws.merge_cells(
        start_row=header_row,
        start_column=OVERTIME_COMP_PREV_REMAIN_COL + 3,
        end_row=header_row + 1,
        end_column=OVERTIME_COMP_PREV_REMAIN_COL + 5,
    )
    _style_overtime_summary_header_merge(
        ws,
        header_row,
        OVERTIME_COMP_PREV_REMAIN_COL + 3,
        header_row + 1,
        OVERTIME_COMP_PREV_REMAIN_COL + 5,
        labels["month_deduct"],
    )

    ws.merge_cells(
        start_row=header_row,
        start_column=OVERTIME_COMP_REMAIN_COL,
        end_row=sub_row,
        end_column=OVERTIME_COMP_REMAIN_COL,
    )
    _style_overtime_summary_header_merge(
        ws,
        header_row,
        OVERTIME_COMP_REMAIN_COL,
        sub_row,
        OVERTIME_COMP_REMAIN_COL,
        labels["month_remain"],
    )

    _style_overtime_header_cell(ws, sub_row, OVERTIME_COMP_PREV_REMAIN_COL, labels["prev_remain"])
    ws.merge_cells(
        start_row=sub_row,
        start_column=OVERTIME_COMP_MONTH_HOURS_COL,
        end_row=sub_row,
        end_column=OVERTIME_COMP_MONTH_HOURS_COL + 1,
    )
    _style_overtime_header_cell(ws, sub_row, OVERTIME_COMP_MONTH_HOURS_COL, "1倍")
    ws.merge_cells(
        start_row=sub_row,
        start_column=OVERTIME_COMP_PREV_REMAIN_COL + 3,
        end_row=sub_row,
        end_column=OVERTIME_COMP_PREV_REMAIN_COL + 5,
    )
    _style_overtime_header_cell(ws, sub_row, OVERTIME_COMP_PREV_REMAIN_COL + 3, "1倍")


def write_overtime_compensatory_formulas(ws, row: int) -> None:
    ws.cell(
        row=row,
        column=OVERTIME_COMP_MONTH_HOURS_COL,
        value=overtime_comp_month_hours_formula(row),
    )
    ws.cell(
        row=row,
        column=OVERTIME_COMP_REMAIN_COL,
        value=overtime_comp_remain_formula(row),
    )
    _apply_overtime_summary_column_style(ws, row)


def write_overtime_compensatory_employee_row(
    ws,
    row: int,
    *,
    name: str = "",
    department: str = "",
) -> None:
    ws.cell(row=row, column=1, value=name or None)
    ws.cell(row=row, column=2, value=department or None)
    ws.cell(row=row, column=3, value=OVERTIME_SETTLEMENT_COMP)
    write_overtime_compensatory_formulas(ws, row)


def write_overtime_compensatory_section(
    ws,
    year: int,
    month: int,
    employees: Sequence[TemplateEmployee],
) -> None:
    """Build the 调休 block (rows 90–186): headers, employee rows, and blank slots."""
    write_overtime_compensatory_headers(ws, year, month)

    slot_count = OVERTIME_COMP_MAX_ROW - OVERTIME_COMP_DATA_START_ROW + 1
    resolved = list(employees)
    if len(resolved) < slot_count:
        resolved.extend(_blank_employees(slot_count - len(resolved)))

    for offset in range(slot_count):
        row = OVERTIME_COMP_DATA_START_ROW + offset
        employee = resolved[offset]
        write_overtime_compensatory_employee_row(
            ws,
            row,
            name=employee.name,
            department=employee.department,
        )

    calc_end_col = OVERTIME_CALC_START_COL + 6
    _apply_border_range(
        ws,
        OVERTIME_COMP_HEADER_ROW,
        OVERTIME_COMP_MAX_ROW,
        1,
        calc_end_col,
    )


def write_overtime_employee_calc_formulas(ws, row: int) -> None:
    """Write AJ–AP calculation formulas for one employee row on 加班结算加班工资."""
    formulas = (
        overtime_15x_hours_formula(row),
        overtime_2x_hours_formula(row),
        overtime_3x_hours_formula(row),
        overtime_15x_total_formula(row),
        overtime_2x_total_formula(row),
        overtime_3x_total_formula(row),
        overtime_pay_total_formula(row),
    )
    for idx, formula in enumerate(formulas):
        cell = ws.cell(row=row, column=OVERTIME_CALC_START_COL + idx, value=formula)
    _apply_overtime_summary_column_style(ws, row)


def write_overtime_calc_formulas(ws, employee_count: int) -> None:
    """Write AJ–AP formulas for every employee on 加班结算加班工资 (from row 5)."""
    for index in range(employee_count):
        row = OVERTIME_DATA_START_ROW + index
        write_overtime_employee_calc_formulas(ws, row)


def _blank_employees(count: int) -> Tuple[TemplateEmployee, ...]:
    return tuple(TemplateEmployee(name="") for _ in range(count))


def _resolve_employees(employees: Optional[Sequence[TemplateEmployee]]) -> Sequence[TemplateEmployee]:
    if not employees:
        return _blank_employees(DEFAULT_BLANK_EMPLOYEE_SLOTS)
    if len(employees) < DEFAULT_BLANK_EMPLOYEE_SLOTS:
        return tuple(employees) + _blank_employees(DEFAULT_BLANK_EMPLOYEE_SLOTS - len(employees))
    return employees


def sign_legend_cell_text(symbol: str, label: str) -> str:
    """Format one 签字 sheet legend cell, e.g. ``√ (出勤)``."""
    return f"{symbol} ({label})"


def write_sign_sheet_legend(ws) -> None:
    """Write symbol legend to AJ5–AS5 on the 签字 sheet.

    AJ–AR store the symbol only (COUNTIF criteria in row 5). AS stores the full label.
    """
    for idx, (symbol, label) in enumerate(SIGN_LEGEND_SYMBOLS):
        col = SIGN_SUMMARY_START_COL + idx
        if idx < len(SIGN_COUNT_SYMBOLS):
            value = symbol
        else:
            value = sign_legend_cell_text(symbol, label)
        cell = ws.cell(row=SIGN_LEGEND_ROW, column=col, value=value)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL


def _build_sign_sheet(ws, year: int, month: int, employees: Sequence[TemplateEmployee]) -> None:
    days_in_month = calendar.monthrange(year, month)[1]
    work_days = count_month_work_days(year, month)
    last_col = max(SIGN_ABSENT_COL, SIGN_SUMMARY_END_COL + 6)

    _merge_title(ws, 1, 1, last_col, f"{COMPANY_NAME}员工{year}年{month}月考勤表")
    ws.row_dimensions[1].height = 28

    for col, label in ((1, "签名"), (2, "姓名"), (3, "时间")):
        cell = ws.cell(row=SIGN_HEADER_ROW, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for day in range(1, SIGN_DAY_COUNT + 1):
        col = SIGN_DAY_START_COL + day - 1
        value = day if day <= days_in_month else f"{day}*"
        cell = ws.cell(row=SIGN_DATE_ROW, column=col, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        if day > days_in_month:
            cell.font = Font(name="宋体", size=10, bold=True, color="999999")

    write_sign_sheet_legend(ws)
    absent_header = ws.cell(row=SIGN_DATE_ROW, column=SIGN_ABSENT_COL, value="缺勤")
    absent_header.font = HEADER_FONT
    absent_header.fill = HEADER_FILL
    absent_header.alignment = CENTER

    current_row = SIGN_DATA_START_ROW
    for employee in employees:
        am_row = current_row
        pm_row = current_row + 1

        ws.merge_cells(start_row=am_row, start_column=1, end_row=pm_row, end_column=1)
        ws.cell(row=am_row, column=1).alignment = CENTER

        ws.merge_cells(start_row=am_row, start_column=2, end_row=pm_row, end_column=2)
        name_cell = ws.cell(row=am_row, column=2, value=employee.name or None)
        name_cell.font = BODY_FONT
        name_cell.alignment = CENTER

        ws.cell(row=am_row, column=3, value="上午").alignment = CENTER
        ws.cell(row=pm_row, column=3, value="下午").alignment = CENTER

        for day in range(1, SIGN_DAY_COUNT + 1):
            col = SIGN_DAY_START_COL + day - 1
            if day > days_in_month:
                for row in (am_row, pm_row):
                    ws.cell(row=row, column=col, value="")
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
            else:
                ws.cell(row=am_row, column=col, value="")
                ws.cell(row=pm_row, column=col, value="")

        for idx, symbol in enumerate(SIGN_COUNT_SYMBOLS):
            col = SIGN_SUMMARY_START_COL + idx
            for row in (am_row, pm_row):
                ws.cell(row=row, column=col, value=sign_countif_formula(row, idx))
                ws.cell(row=row, column=col).alignment = CENTER

        for row in (am_row, pm_row):
            ws.cell(
                row=row,
                column=SIGN_ABSENT_COL,
                value=sign_absent_formula(row, work_days),
            )
            ws.cell(row=row, column=SIGN_ABSENT_COL).alignment = CENTER

        current_row += 2

    last_data_row = max(current_row - 1, SIGN_LEGEND_ROW)
    _apply_border_range(ws, SIGN_HEADER_ROW, last_data_row, 1, SIGN_ABSENT_COL)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    for col in range(SIGN_DAY_START_COL, SIGN_ABSENT_COL + 1):
        ws.column_dimensions[_col_letter(col)].width = 4


def _build_situation_sheet(ws) -> None:
    headers = ("姓名", "日期", "异常情况", "是否补单", "备注")
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[_col_letter(idx)].width = 18 if idx == 3 else 14

    for row in range(SITUATION_DATA_START_ROW, SITUATION_DATA_START_ROW + 50):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col, value="")
            ws.cell(row=row, column=col).border = THIN_BORDER


def _build_monthly_summary_sheet(
    ws,
    year: int,
    month: int,
    employees: Sequence[TemplateEmployee],
) -> None:
    days_in_month = calendar.monthrange(year, month)[1]

    configure_monthly_summary_headers(ws, year, month)

    for offset, employee in enumerate(employees):
        row = MONTHLY_DATA_START_ROW + offset
        ws.cell(row=row, column=1, value=employee.name or None)

        for day in range(1, SIGN_DAY_COUNT + 1):
            col = MONTHLY_DAILY_START_COL + day - 1
            ws.cell(row=row, column=col, value="" if day > days_in_month else "")

    last_row = MONTHLY_DATA_START_ROW + len(employees) - 1
    format_monthly_summary_sheet(ws, year, month, last_row=last_row)


def overtime_sheet_title(year: int, month: int) -> str:
    return f"{year}年{month}月加班统计汇总表"


def write_overtime_sheet_title(ws, year: int, month: int) -> None:
    """Write row 1 title on 加班结算加班工资."""
    day_end_col = OVERTIME_DAY_START_COL + OVERTIME_DAY_COUNT - 1
    calc_end_col = OVERTIME_CALC_START_COL + 6
    last_col = max(day_end_col, calc_end_col)
    _merge_title(ws, OVERTIME_TITLE_ROW, 1, last_col, overtime_sheet_title(year, month))


def _build_overtime_sheet(ws, year: int, month: int, employees: Sequence[TemplateEmployee]) -> None:
    day_end_col = OVERTIME_DAY_START_COL + OVERTIME_DAY_COUNT - 1
    calc_end_col = OVERTIME_CALC_END_COL

    write_overtime_sheet_title(ws, year, month)
    write_overtime_pay_section_headers(ws, year, month)

    days_in_month = calendar.monthrange(year, month)[1]
    for offset, employee in enumerate(employees):
        row = OVERTIME_DATA_START_ROW + offset
        ws.cell(row=row, column=1, value=employee.name or None)
        ws.cell(row=row, column=2, value=employee.department)
        ws.cell(row=row, column=3, value=OVERTIME_SETTLEMENT_PAY)

        for day in range(1, OVERTIME_DAY_COUNT + 1):
            col = OVERTIME_DAY_START_COL + day - 1
            if day > days_in_month:
                ws.cell(row=row, column=col, value="")
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
            else:
                ws.cell(row=row, column=col, value=0)

        write_overtime_employee_calc_formulas(ws, row)

    last_row = OVERTIME_DATA_START_ROW + len(employees) - 1
    _apply_border_range(ws, OVERTIME_MAIN_HEADER_ROW, last_row, 1, calc_end_col)
    write_overtime_compensatory_section(ws, year, month, employees)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    for col in range(OVERTIME_DAY_START_COL, day_end_col + 1):
        ws.column_dimensions[_col_letter(col)].width = 4


def build_attendance_workbook(
    year: int,
    month: int,
    employees: Optional[Sequence[TemplateEmployee]] = None,
) -> Workbook:
    resolved = _resolve_employees(employees)
    workbook = Workbook()
    sign_ws = workbook.active
    sign_ws.title = TEMPLATE_SHEETS[0]

    situation_ws = workbook.create_sheet(TEMPLATE_SHEETS[1])
    monthly_ws = workbook.create_sheet(TEMPLATE_SHEETS[2])
    overtime_ws = workbook.create_sheet(TEMPLATE_SHEETS[3])

    _build_sign_sheet(sign_ws, year, month, resolved)
    _build_situation_sheet(situation_ws)
    _build_monthly_summary_sheet(monthly_ws, year, month, resolved)
    _build_overtime_sheet(overtime_ws, year, month, resolved)
    return workbook


def generate_attendance_template(
    output_path: Union[str, Path],
    year: int,
    month: int,
    employees: Optional[Sequence[TemplateEmployee]] = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_attendance_workbook(year, month, employees)
    workbook.save(path)
    return path
