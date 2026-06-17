"""
DingTalk-style daily status text and color fills for the 月度汇总 sheet.

The monthly summary shows descriptive sync text (e.g. ``出差05-06 08:30到05-16 17:30 11天``,
``上班迟到9分钟``) — never symbol checkmarks used on the 签字 sheet.
"""

from __future__ import annotations

import calendar
from typing import Optional, Tuple

from openpyxl.styles import Font, PatternFill

from app.models import MonthlyAttendance

# Legacy symbol / short codes stored before full DingTalk text sync.
LEGACY_SYMBOL_TO_MONTHLY_TEXT = {
    "√": "正常",
    "▼": "出差",
    "◇": "事假",
    "✬": "调休",
    "※": "病假",
    "●": "福利假",
    "AL": "年假",
    "○": "产假",
    "FL": "丧假",
    "ML": "婚假",
    "旷工": "旷工",
    "旷": "旷工",
    "迟到": "上班迟到",
    "迟": "上班迟到",
    "缺卡": "上班缺卡",
    "缺": "上班缺卡",
    "未打卡": "未打卡",
    "休息": "休息",
    "正常": "正常",
    "出勤": "正常",
    "present": "正常",
    "business_trip": "出差",
    "personal_leave": "事假",
    "compensatory_leave": "调休",
    "sick_leave": "病假",
    "welfare_leave": "福利假",
    "annual_leave": "年假",
    "maternity_leave": "产假",
    "funeral_leave": "丧假",
    "marriage_leave": "婚假",
    "absenteeism": "旷工",
    "late": "上班迟到",
    "missing_punch": "上班缺卡",
    "rest": "休息",
}

HALF_DAY_SEPARATORS = ("/", "|", "、")

MONTHLY_FILL_CHUCHAI = PatternFill("solid", fgColor="FFA500")
MONTHLY_FILL_KUANGGONG = PatternFill("solid", fgColor="FFA6D2")
MONTHLY_FILL_QUEKA = PatternFill("solid", fgColor="FF0000")
MONTHLY_FILL_CHIDAO = PatternFill("solid", fgColor="90EE90")
MONTHLY_FILL_WAICHU = PatternFill("solid", fgColor="ADD8E6")
MONTHLY_FILL_BINGJIA = PatternFill("solid", fgColor="0000FF")
MONTHLY_FILL_SHIJIA = PatternFill("solid", fgColor="FFFF00")
MONTHLY_FILL_XIUXI = PatternFill("solid", fgColor="E8E8E8")
MONTHLY_FILL_JIABAN = PatternFill("solid", fgColor="800080")

MONTHLY_FONT_ON_DARK = Font(color="FFFFFF", bold=False)
MONTHLY_FONT_ON_LIGHT = Font(color="000000", bold=False)


def map_monthly_summary_text(raw: str) -> str:
    """Convert a single status token to DingTalk-style monthly summary text."""
    if not raw:
        return ""
    text = raw.strip()
    if not text:
        return ""
    if text in LEGACY_SYMBOL_TO_MONTHLY_TEXT:
        return LEGACY_SYMBOL_TO_MONTHLY_TEXT[text]
    return text


def _is_half_day_mark(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    if any(separator in text for separator in HALF_DAY_SEPARATORS):
        return True
    single_char_symbols = {"√", "◇", "✬", "▼", "※", "●", "○"}
    return (
        len(text) == 2
        and text[0] in single_char_symbols
        and text[1] in single_char_symbols
    )


def _split_day_halves(value: str) -> Tuple[str, str]:
    text = value.strip()
    if not _is_half_day_mark(text):
        return text, text
    for separator in HALF_DAY_SEPARATORS:
        if separator in text:
            left, right = text.split(separator, 1)
            return left.strip(), right.strip()
    if len(text) == 2:
        return text[0], text[1]
    return text, text


def combined_monthly_day_status(am_value: str, pm_value: str) -> str:
    mapped_am = map_monthly_summary_text(am_value)
    mapped_pm = map_monthly_summary_text(pm_value)
    if mapped_am and mapped_pm and mapped_am == mapped_pm:
        return mapped_am
    if mapped_am and mapped_pm:
        return f"{mapped_am},{mapped_pm}"
    return mapped_am or mapped_pm


def format_monthly_summary_day_status(
    record: MonthlyAttendance,
    year: int,
    month: int,
    day: int,
    *,
    resolve_day_value,
) -> Optional[str]:
    """Return descriptive monthly-summary text for ``day_{day}`` (never ``√``)."""
    days_in_month = calendar.monthrange(year, month)[1]
    if day > days_in_month:
        return None

    day_value = resolve_day_value(record, day)
    if not day_value:
        return None

    text = day_value.strip()
    if not text:
        return None

    if not _is_half_day_mark(text):
        display = map_monthly_summary_text(text)
        return display if display else None

    am_value, pm_value = _split_day_halves(text)
    display = combined_monthly_day_status(am_value, pm_value)
    return display if display else None


def monthly_summary_status_style(text: str) -> Tuple[Optional[PatternFill], Optional[Font]]:
    """Return background fill and optional font for a monthly-summary cell value."""
    if not text:
        return None, None

    value = text.strip()
    if not value:
        return None, None

    if "出差" in value:
        return MONTHLY_FILL_CHUCHAI, MONTHLY_FONT_ON_LIGHT
    if "旷工" in value:
        return MONTHLY_FILL_KUANGGONG, MONTHLY_FONT_ON_LIGHT
    if "缺卡" in value or value == "未打卡":
        return MONTHLY_FILL_QUEKA, MONTHLY_FONT_ON_DARK
    if "迟到" in value or "早退" in value:
        return MONTHLY_FILL_CHIDAO, MONTHLY_FONT_ON_LIGHT
    if "外勤" in value or "外出" in value:
        return MONTHLY_FILL_WAICHU, MONTHLY_FONT_ON_LIGHT
    if "病假" in value or value.startswith("※"):
        return MONTHLY_FILL_BINGJIA, MONTHLY_FONT_ON_DARK
    if "事假" in value:
        return MONTHLY_FILL_SHIJIA, MONTHLY_FONT_ON_LIGHT
    if value == "休息" or value.startswith("休息"):
        return MONTHLY_FILL_XIUXI, MONTHLY_FONT_ON_LIGHT
    if "加班" in value:
        return MONTHLY_FILL_JIABAN, MONTHLY_FONT_ON_DARK

    return None, None
