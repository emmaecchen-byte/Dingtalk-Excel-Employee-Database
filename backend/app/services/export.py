"""
Export attendance period data to Excel and PDF (spec section 6.6).

Excel: 2-sheet workbook (签字 + 情况说明) with formulas.
PDF: see ``app.services.pdf_export`` (landscape attendance grid + exception table).
"""

from __future__ import annotations

import calendar
import logging
import tempfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.crud.abnormal_record import abnormal_record
from app.excel.template_generator import (
    BODY_FONT,
    CENTER,
    HEADER_FILL,
    HEADER_FONT,
    SIGN_ABSENT_COL,
    SIGN_DATA_START_ROW,
    SIGN_DAY_COUNT,
    SIGN_DAY_START_COL,
    SIGN_HEADER_ROW,
    SIGN_LEGEND_ROW,
    SITUATION_DATA_START_ROW,
    THIN_BORDER,
    _apply_border_range,
    count_month_work_days,
    is_calendar_weekend,
    write_sign_sheet_employee_am_pm_summary_formulas,
    write_sign_sheet_headers,
    write_sign_sheet_legend,
)
from app.models import AbnormalRecord, AttendancePeriod, EmployeeAttendance
from app.services.attendance_rule_engine import load_company_rules
from app.services.exception_detection import EXCEPTION_LABELS

logger = logging.getLogger(__name__)

SITUATION_HEADERS = ("姓名", "日期", "异常情况", "是否补单", "备注")
SITUATION_COL_COUNT = len(SITUATION_HEADERS)

SUPPLEMENT_EXPORT_LABELS = {
    "yes": "是",
    "no": "否",
    "pending": "待处理",
    "not_required": "不需要",
}

OUT_OF_MONTH_FILL = PatternFill("solid", fgColor="F2F2F2")
WEEKEND_FILL = PatternFill("solid", fgColor="D9D9D9")
SUMMARY_ROW_FILL = PatternFill("solid", fgColor="E2EFDA")


class PeriodExportError(Exception):
    def __init__(self, message: str, status_code: int = 404):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


@dataclass
class PeriodExcelExportResult:
    path: Path
    filename: str
    period_id: int
    year: int
    month: int
    employee_count: int
    exception_count: int

    def as_stream(self) -> BytesIO:
        buffer = BytesIO(self.path.read_bytes())
        buffer.seek(0)
        return buffer

    def cleanup(self) -> None:
        try:
            self.path.unlink(missing_ok=True)
        except OSError:
            logger.warning("Failed to remove temporary export file: %s", self.path)


def _load_period(
    db: Session,
    period_id: int,
    company_id: int,
) -> AttendancePeriod:
    period = (
        db.query(AttendancePeriod)
        .filter(AttendancePeriod.id == period_id, AttendancePeriod.company_id == company_id)
        .first()
    )
    if not period:
        raise PeriodExportError("Attendance period not found", status_code=404)
    return period


def _load_employees(db: Session, period_id: int) -> List[EmployeeAttendance]:
    return (
        db.query(EmployeeAttendance)
        .options(joinedload(EmployeeAttendance.daily_records))
        .filter(EmployeeAttendance.period_id == period_id)
        .order_by(EmployeeAttendance.row_index, EmployeeAttendance.id)
        .all()
    )


def _half_day_symbol(status: Optional[str], rules) -> str:
    from app.services.attendance_rule_engine import map_status_to_symbol

    return map_status_to_symbol(status or "", rules)


def _write_sign_day_cell(ws, row: int, col: int, symbol: str, *, year: int, month: int, day: int) -> None:
    cell = ws.cell(row=row, column=col, value=symbol or None)
    cell.alignment = CENTER
    cell.font = BODY_FONT
    if symbol:
        return
    days_in_month = calendar.monthrange(year, month)[1]
    if day > days_in_month:
        cell.fill = OUT_OF_MONTH_FILL
    elif is_calendar_weekend(year, month, day):
        cell.fill = WEEKEND_FILL


def _populate_sign_sheet(
    ws,
    *,
    year: int,
    month: int,
    employees: Sequence[EmployeeAttendance],
    rules,
) -> int:
    """Write 签字 sheet from period daily records. Returns last data row."""
    days_in_month = calendar.monthrange(year, month)[1]
    write_sign_sheet_headers(ws, year, month)
    write_sign_sheet_legend(ws)

    for index, employee_row in enumerate(employees):
        am_row = SIGN_DATA_START_ROW + index * 2
        pm_row = am_row + 1
        daily_by_day = {item.day: item for item in employee_row.daily_records}
        name = employee_row.employee_name or ""

        for row in (am_row, pm_row):
            ws.cell(row=row, column=1, value=None)
            name_cell = ws.cell(row=row, column=2, value=name or None)
            name_cell.font = BODY_FONT
            name_cell.alignment = CENTER

        ws.cell(row=am_row, column=3, value="上午").alignment = CENTER
        ws.cell(row=pm_row, column=3, value="下午").alignment = CENTER

        for day in range(1, SIGN_DAY_COUNT + 1):
            col = SIGN_DAY_START_COL + day - 1
            record = daily_by_day.get(day)
            if day > days_in_month:
                _write_sign_day_cell(ws, am_row, col, "", year=year, month=month, day=day)
                _write_sign_day_cell(ws, pm_row, col, "", year=year, month=month, day=day)
                continue

            am_symbol = _half_day_symbol(record.morning_status if record else None, rules)
            pm_symbol = _half_day_symbol(record.afternoon_status if record else None, rules)
            _write_sign_day_cell(ws, am_row, col, am_symbol, year=year, month=month, day=day)
            _write_sign_day_cell(ws, pm_row, col, pm_symbol, year=year, month=month, day=day)

        write_sign_sheet_employee_am_pm_summary_formulas(
            ws,
            am_row,
            pm_row,
            work_days=count_month_work_days(year, month),
        )

    last_row = SIGN_DATA_START_ROW + len(employees) * 2 - 1 if employees else SIGN_LEGEND_ROW
    _apply_border_range(ws, SIGN_HEADER_ROW, last_row, 1, SIGN_ABSENT_COL)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    for col in range(SIGN_DAY_START_COL, SIGN_ABSENT_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4

    ws.freeze_panes = ws.cell(row=SIGN_DATA_START_ROW, column=SIGN_DAY_START_COL).coordinate
    return last_row


def _format_dates_column(record: AbnormalRecord) -> str:
    dates = record.dates or []
    if dates:
        return "、".join(item.get("date", "") for item in dates if item.get("date"))
    return ""


def _supplement_label(status: str) -> str:
    return SUPPLEMENT_EXPORT_LABELS.get(status, status)


def _write_situation_headers(ws) -> None:
    for idx, header in enumerate(SITUATION_HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        width = 22 if idx == 3 else 16 if idx in (2, 5) else 12
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_situation_data_row(ws, row: int, record: AbnormalRecord) -> None:
    values = (
        record.employee_name,
        _format_dates_column(record) or record.summary,
        record.summary or EXCEPTION_LABELS.get(record.exception_type, record.exception_type),
        _supplement_label(record.supplement_status),
        record.notes or "",
    )
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value or None)
        cell.font = BODY_FONT
        cell.alignment = Alignment(
            horizontal="left" if col in (3, 5) else "center",
            vertical="center",
            wrap_text=col in (3, 5),
        )
        cell.border = THIN_BORDER


def _write_situation_summary_formulas(ws, *, first_data_row: int, last_data_row: int) -> int:
    """Footer row with COUNT / COUNTIF formulas (not hard-coded totals)."""
    if last_data_row < first_data_row:
        summary_row = first_data_row + 1
        ws.cell(row=summary_row, column=1, value="异常合计")
        ws.cell(row=summary_row, column=3, value=0)
        return summary_row

    summary_row = last_data_row + 2
    data_range = f"A{first_data_row}:A{last_data_row}"
    supplement_range = f"D{first_data_row}:D{last_data_row}"

    labels = (
        ("异常合计", f"=COUNTA({data_range})"),
        ("待补单", f'=COUNTIF({supplement_range},"待处理")'),
        ("已补单", f'=COUNTIF({supplement_range},"是")'),
    )
    for offset, (label, formula) in enumerate(labels):
        row = summary_row + offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name="宋体", size=10, bold=True)
        label_cell.fill = SUMMARY_ROW_FILL
        label_cell.border = THIN_BORDER

        value_cell = ws.cell(row=row, column=3, value=formula)
        value_cell.font = Font(name="宋体", size=10, bold=True)
        value_cell.alignment = CENTER
        value_cell.fill = SUMMARY_ROW_FILL
        value_cell.border = THIN_BORDER

        for col in (2, 4, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = SUMMARY_ROW_FILL
            cell.border = THIN_BORDER

    return summary_row + len(labels) - 1


def _populate_situation_sheet(ws, records: Sequence[AbnormalRecord]) -> int:
    """Write 情况说明 sheet. Returns last row including summary block."""
    _write_situation_headers(ws)

    row = SITUATION_DATA_START_ROW
    for record in records:
        _write_situation_data_row(ws, row, record)
        row += 1

    last_data_row = row - 1 if records else SITUATION_DATA_START_ROW - 1
    if last_data_row >= SITUATION_DATA_START_ROW:
        ws.auto_filter.ref = f"A1:{get_column_letter(SITUATION_COL_COUNT)}{last_data_row}"

    return _write_situation_summary_formulas(
        ws,
        first_data_row=SITUATION_DATA_START_ROW,
        last_data_row=last_data_row,
    )


def build_period_workbook(
    *,
    period: AttendancePeriod,
    employees: Sequence[EmployeeAttendance],
    exceptions: Sequence[AbnormalRecord],
    rules,
) -> Workbook:
    workbook = Workbook()
    sign_ws = workbook.active
    sign_ws.title = "签字"
    situation_ws = workbook.create_sheet("情况说明")

    _populate_sign_sheet(sign_ws, year=period.year, month=period.month, employees=employees, rules=rules)
    _populate_situation_sheet(situation_ws, exceptions)
    return workbook


def generate_period_excel(
    db: Session,
    period_id: int,
    company_id: int,
) -> PeriodExcelExportResult:
    period = _load_period(db, period_id, company_id)
    employees = _load_employees(db, period_id)
    if not employees:
        raise PeriodExportError("No attendance data found for this period", status_code=404)

    exceptions = abnormal_record.list_for_period(
        db,
        period_id=period_id,
        company_id=company_id,
    )

    rules = load_company_rules(db, company_id)
    workbook = build_period_workbook(period=period, employees=employees, exceptions=exceptions, rules=rules)

    filename = f"attendance_period_{period.year}_{period.month:02d}_{period_id}.xlsx"
    temp_dir = Path(tempfile.gettempdir())
    path = temp_dir / f"period_export_{period_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    workbook.save(path)

    logger.info(
        "Period Excel export generated period_id=%s company_id=%s employees=%s exceptions=%s file=%s",
        period_id,
        company_id,
        len(employees),
        len(exceptions),
        path,
    )

    return PeriodExcelExportResult(
        path=path,
        filename=filename,
        period_id=period_id,
        year=period.year,
        month=period.month,
        employee_count=len(employees),
        exception_count=len(exceptions),
    )
