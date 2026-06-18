"""
DingTalk monthly summary Excel parser and validation (spec sections 4.1, 7.3, 12).
"""

from __future__ import annotations

import calendar
import logging
import re
from dataclasses import dataclass, field
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, List, Optional, Sequence, Tuple, Union

from openpyxl import load_workbook

from app.excel.field_utils import normalize_value
from app.excel.template_generator import (
    MONTHLY_DAILY_START_COL,
    MONTHLY_DATA_START_ROW,
    MONTHLY_TITLE_ROW,
    TEMPLATE_SHEETS,
)
from app.services.attendance_rule_engine import (
    ResolvedAttendanceRule,
    default_rules,
    is_known_status as rule_is_known_status,
)

logger = logging.getLogger(__name__)

MULTI_STATUS_SPLIT_RE = re.compile(r"\s*\+\s*|[、;；]")
HALF_DAY_SPLIT_RE = re.compile(r"[/|]")


class ValidationSeverity(str, Enum):
    ERROR = "Error"
    WARNING = "Warning"
    INFO = "Info"


KNOWN_STATUS_KEYWORDS = (
    "正常",
    "出勤",
    "缺卡",
    "未打卡",
    "迟到",
    "严重迟到",
    "旷工",
    "出差",
    "事假",
    "病假",
    "年假",
    "调休",
    "产假",
    "陪产假",
    "婚假",
    "丧假",
    "福利假",
    "休息",
    "请假",
    "补卡",
    "外勤",
    "加班",
    "上班",
    "下班",
)

MODIFIER_KEYWORDS = ("补卡", "补卡申请", "申请")


class DingTalkParserError(Exception):
    def __init__(self, message: str, status_code: int = 400):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


@dataclass
class ValidationIssue:
    severity: ValidationSeverity
    code: str
    message: str
    employee_name: Optional[str] = None
    day: Optional[int] = None
    row_index: Optional[int] = None

    def to_dict(self) -> dict:
        return {
            "severity": self.severity.value,
            "code": self.code,
            "message": self.message,
            "employee_name": self.employee_name,
            "day": self.day,
            "row_index": self.row_index,
        }


@dataclass
class ParsedDailyCell:
    day: int
    raw_text: str
    morning_status: Optional[str] = None
    afternoon_status: Optional[str] = None
    requires_review: bool = False


@dataclass
class ParsedEmployeeAttendance:
    employee_name: str
    row_index: int
    daily_cells: List[ParsedDailyCell] = field(default_factory=list)
    requires_review: bool = False


@dataclass
class ParsedDingTalkWorkbook:
    year: int
    month: int
    sheet_name: str
    employees: List[ParsedEmployeeAttendance]
    issues: List[ValidationIssue] = field(default_factory=list)

    @property
    def has_blocking_errors(self) -> bool:
        return any(issue.severity == ValidationSeverity.ERROR for issue in self.issues)

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == ValidationSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == ValidationSeverity.WARNING)

    @property
    def info_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == ValidationSeverity.INFO)


TITLE_YEAR_MONTH_RE = re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月")


def _parse_title_period(title: object) -> Tuple[Optional[int], Optional[int]]:
    if not title:
        return None, None
    text = str(title).strip()
    match = TITLE_YEAR_MONTH_RE.search(text)
    if match:
        return int(match.group(1)), int(match.group(2))
    if "统计日期" in text and "至" in text:
        try:
            start = text.split("至", 1)[0]
            year_month = start.split("：", 1)[-1].strip()[:7]
            parsed_year, parsed_month = year_month.split("-", 1)
            return int(parsed_year), int(parsed_month)
        except (TypeError, ValueError, IndexError):
            return None, None
    return None, None


def _find_monthly_sheet(wb) -> Tuple[object, str]:
    if TEMPLATE_SHEETS[2] in wb.sheetnames:
        return wb[TEMPLATE_SHEETS[2]], TEMPLATE_SHEETS[2]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        title = ws.cell(row=MONTHLY_TITLE_ROW, column=1).value
        if title and ("月度汇总" in str(title) or "统计日期" in str(title)):
            return ws, sheet_name

    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]], wb.sheetnames[0]

    raise DingTalkParserError(
        'Unable to locate monthly summary sheet. Expected worksheet named "月度汇总".'
    )


def _is_known_status(text: str, rules: Sequence[ResolvedAttendanceRule]) -> bool:
    return rule_is_known_status(text, rules)


def _is_modifier_only(text: str, rules: Sequence[ResolvedAttendanceRule]) -> bool:
    normalized = text.strip()
    return any(keyword in normalized for keyword in MODIFIER_KEYWORDS) and not _is_known_status(
        normalized, rules
    )


def _split_cell_statuses(
    raw_text: str,
    rules: Sequence[ResolvedAttendanceRule],
) -> Tuple[str, Optional[str], Optional[str], bool]:
    """
    Parse a daily cell into morning/afternoon statuses.

    Handles:
    - single status: 正常
    - half-day split: 正常/迟到
    - compound status: 正常 + 补卡申请
    """
    text = (raw_text or "").strip()
    if not text:
        return "", None, None, False

    requires_review = False

    if HALF_DAY_SPLIT_RE.search(text):
        parts = [part.strip() for part in HALF_DAY_SPLIT_RE.split(text, maxsplit=1) if part.strip()]
        if len(parts) == 2:
            morning, afternoon = parts
            if not _is_known_status(morning, rules) or not _is_known_status(afternoon, rules):
                requires_review = True
            return text, morning, afternoon, requires_review
        if len(parts) == 1:
            value = parts[0]
            if not _is_known_status(value, rules):
                requires_review = True
            return text, value, value, requires_review

    if MULTI_STATUS_SPLIT_RE.search(text):
        parts = [part.strip() for part in MULTI_STATUS_SPLIT_RE.split(text) if part.strip()]
        primary_parts = [part for part in parts if not _is_modifier_only(part, rules)]
        if not primary_parts:
            primary_parts = parts[:1]
        primary = primary_parts[0]
        if len(primary_parts) > 1:
            requires_review = True
        if any(not _is_known_status(part, rules) for part in parts):
            requires_review = True
        if not _is_known_status(primary, rules):
            requires_review = True
        return text, primary, primary, requires_review

    if not _is_known_status(text, rules):
        requires_review = True
    return text, text, text, requires_review


def _detect_data_start_row(ws) -> int:
    for row_index in range(1, 15):
        header = ws.cell(row=row_index, column=1).value
        if header is not None and str(header).strip() == "姓名":
            return row_index + 2
        day_header = ws.cell(row=row_index, column=MONTHLY_DAILY_START_COL).value
        if day_header is not None and str(day_header).strip() in {"1", "01"}:
            return row_index + 1
    return MONTHLY_DATA_START_ROW


def _parse_employee_rows(
    ws,
    year: int,
    month: int,
    rules: Sequence[ResolvedAttendanceRule],
) -> List[ParsedEmployeeAttendance]:
    days_in_month = calendar.monthrange(year, month)[1]
    data_start_row = _detect_data_start_row(ws)
    employees: List[ParsedEmployeeAttendance] = []

    row_index = data_start_row
    while True:
        name_value = ws.cell(row=row_index, column=1).value
        if name_value is None or str(name_value).strip() == "":
            break

        employee_name = str(name_value).strip()
        daily_cells: List[ParsedDailyCell] = []

        for day in range(1, days_in_month + 1):
            col = MONTHLY_DAILY_START_COL + day - 1
            raw_text = normalize_value(ws.cell(row=row_index, column=col).value)
            raw_text, morning, afternoon, requires_review = _split_cell_statuses(raw_text, rules)
            daily_cells.append(
                ParsedDailyCell(
                    day=day,
                    raw_text=raw_text,
                    morning_status=morning,
                    afternoon_status=afternoon,
                    requires_review=requires_review,
                )
            )

        employee_requires_review = any(cell.requires_review for cell in daily_cells)
        employees.append(
            ParsedEmployeeAttendance(
                employee_name=employee_name,
                row_index=row_index,
                daily_cells=daily_cells,
                requires_review=employee_requires_review,
            )
        )
        row_index += 1

    return employees


def validate_parsed_workbook(parsed: ParsedDingTalkWorkbook) -> List[ValidationIssue]:
    issues: List[ValidationIssue] = list(parsed.issues)
    days_in_month = calendar.monthrange(parsed.year, parsed.month)[1]

    if not parsed.employees:
        issues.append(
            ValidationIssue(
                severity=ValidationSeverity.ERROR,
                code="no_employees",
                message="No employee rows found in the uploaded workbook.",
            )
        )
        return issues

    seen_names: Dict[str, int] = {}
    for employee in parsed.employees:
        name = employee.employee_name.strip()
        if not name:
            issues.append(
                ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    code="empty_employee_name",
                    message="Employee name is empty.",
                    row_index=employee.row_index,
                )
            )
            continue

        if name in seen_names:
            issues.append(
                ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    code="duplicate_employee_name",
                    message=f'Duplicate employee name "{name}".',
                    employee_name=name,
                    row_index=employee.row_index,
                )
            )
        seen_names[name] = employee.row_index

        populated_days = [cell.day for cell in employee.daily_cells if cell.raw_text]
        if len(populated_days) < days_in_month:
            issues.append(
                ValidationIssue(
                    severity=ValidationSeverity.WARNING,
                    code="incomplete_month_coverage",
                    message=(
                        f'Employee "{name}" is missing attendance values for '
                        f"{days_in_month - len(populated_days)} day(s)."
                    ),
                    employee_name=name,
                    row_index=employee.row_index,
                )
            )

        for cell in employee.daily_cells:
            if not cell.raw_text:
                continue
            if cell.requires_review:
                issues.append(
                    ValidationIssue(
                        severity=ValidationSeverity.INFO,
                        code="unrecognized_status",
                        message=(
                            f'Unrecognized or compound status "{cell.raw_text}" requires HR review.'
                        ),
                        employee_name=name,
                        day=cell.day,
                        row_index=employee.row_index,
                    )
                )

    return issues


def parse_dingtalk_workbook(
    source: Union[str, Path, BinaryIO, BytesIO],
    *,
    fallback_year: Optional[int] = None,
    fallback_month: Optional[int] = None,
    attendance_rules: Optional[Sequence[ResolvedAttendanceRule]] = None,
) -> ParsedDingTalkWorkbook:
    rules = list(attendance_rules) if attendance_rules is not None else default_rules()
    try:
        wb = load_workbook(source, read_only=False, data_only=True)
    except Exception as exc:
        raise DingTalkParserError(
            "Unable to open Excel file. The workbook may be corrupt or not a valid .xlsx file."
        ) from exc

    try:
        ws, sheet_name = _find_monthly_sheet(wb)
        inferred_year, inferred_month = _parse_title_period(ws.cell(row=MONTHLY_TITLE_ROW, column=1).value)
        year = inferred_year or fallback_year
        month = inferred_month or fallback_month

        if year is None or month is None:
            raise DingTalkParserError(
                "Could not determine reporting month from workbook header. "
                "Expected title like 月度汇总2026年5月 or 统计日期."
            )
        if month < 1 or month > 12:
            raise DingTalkParserError("Invalid month inferred from workbook header.")

        employees = _parse_employee_rows(ws, year, month, rules)
        parsed = ParsedDingTalkWorkbook(
            year=year,
            month=month,
            sheet_name=sheet_name,
            employees=employees,
        )
        parsed.issues = validate_parsed_workbook(parsed)
        logger.info(
            "Parsed DingTalk workbook sheet=%s period=%s-%02d employees=%s issues=%s",
            sheet_name,
            year,
            month,
            len(employees),
            len(parsed.issues),
        )
        return parsed
    finally:
        wb.close()
