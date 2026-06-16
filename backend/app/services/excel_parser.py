"""
Parse uploaded attendance Excel workbooks (openpyxl).

Reads Sheet 3 ``月度汇总`` for employee names, daily status (AJ–BN), and metadata.
"""

from __future__ import annotations

import calendar
import logging
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook

from app.excel.field_utils import normalize_value
from app.excel.template_generator import (
    MONTHLY_DAILY_START_COL,
    MONTHLY_DATA_START_ROW,
    MONTHLY_TITLE_ROW,
    SIGN_DAY_COUNT,
    TEMPLATE_SHEETS,
)

logger = logging.getLogger(__name__)


class ExcelParserError(Exception):
    def __init__(self, message: str, status_code: int = 400):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


@dataclass
class ParsedEmployeeRow:
    name: str
    supplement_submitted: str = ""
    notes: str = ""
    daily_status: Dict[str, str] = field(default_factory=dict)

    def as_comparison_payload(self) -> Dict[str, object]:
        return {
            "employee_name": self.name,
            "supplement_submitted": self.supplement_submitted,
            "notes": self.notes,
            "daily_status": dict(self.daily_status),
        }


@dataclass
class ParsedWorkbook:
    year: Optional[int]
    month: Optional[int]
    employees: List[ParsedEmployeeRow]

    def employees_by_name(self) -> Dict[str, ParsedEmployeeRow]:
        return {row.name: row for row in self.employees if row.name}


def _cell_value(value) -> str:
    return normalize_value(value)


def _parse_title_period(title: object) -> Tuple[Optional[int], Optional[int]]:
    if not title:
        return None, None
    text = str(title)
    if "年" in text and "月" in text:
        try:
            parsed_year = int(text.split("年", 1)[0].split()[-1])
            parsed_month = int(text.split("年", 1)[1].split("月", 1)[0])
            return parsed_year, parsed_month
        except (TypeError, ValueError, IndexError):
            pass
    if "统计日期" in text and "至" in text:
        try:
            start = text.split("至", 1)[0]
            year_month = start.split("：", 1)[-1].strip()[:7]
            parsed_year, parsed_month = year_month.split("-", 1)
            return int(parsed_year), int(parsed_month)
        except (TypeError, ValueError, IndexError):
            return None, None
    return None, None


def parse_monthly_summary_sheet(ws, year: int, month: int) -> Tuple[Tuple[Optional[int], Optional[int]], List[ParsedEmployeeRow]]:
    """Extract employee rows from the 月度汇总 worksheet."""
    days_in_month = calendar.monthrange(year, month)[1]
    employees: List[ParsedEmployeeRow] = []
    inferred_period: Tuple[Optional[int], Optional[int]] = (None, None)

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_index == MONTHLY_TITLE_ROW and row:
            inferred_period = _parse_title_period(row[0])
        if row_index < MONTHLY_DATA_START_ROW:
            continue
        if not row or not row[0]:
            break

        name = str(row[0]).strip()
        if not name:
            break

        parsed = ParsedEmployeeRow(
            name=name,
            supplement_submitted="",
            notes="",
        )

        for day in range(1, SIGN_DAY_COUNT + 1):
            if day > days_in_month:
                continue
            col_index = MONTHLY_DAILY_START_COL - 1 + day - 1
            if col_index < len(row):
                parsed.daily_status[f"day_{day}"] = _cell_value(row[col_index])

        employees.append(parsed)

    return inferred_period, employees


def parse_uploaded_workbook(
    source: Union[str, Path, BinaryIO, BytesIO],
    *,
    year: int,
    month: int,
) -> ParsedWorkbook:
    """
    Parse an uploaded .xlsx workbook and return employee daily status from 月度汇总.
    """
    if month < 1 or month > 12:
        raise ExcelParserError("Month must be between 1 and 12")

    try:
        wb = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParserError(
            "Unable to open Excel file. The workbook may be corrupt or not a valid .xlsx file."
        ) from exc

    try:
        if TEMPLATE_SHEETS[2] not in wb.sheetnames:
            raise ExcelParserError('Uploaded file is missing the "月度汇总" worksheet')

        monthly_ws = wb[TEMPLATE_SHEETS[2]]
        inferred_period, employees = parse_monthly_summary_sheet(monthly_ws, year, month)

        if not employees:
            raise ExcelParserError("No employee rows found in the 月度汇总 worksheet")

        return ParsedWorkbook(
            year=inferred_period[0] or year,
            month=inferred_period[1] or month,
            employees=employees,
        )
    finally:
        wb.close()
