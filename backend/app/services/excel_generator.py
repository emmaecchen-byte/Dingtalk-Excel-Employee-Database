"""
Excel workbook generation from monthly_attendance records.

Loads ``backend/templates/master_template.xlsx`` with openpyxl and populates:
  - Sheet 1 签字: two rows per employee (上午 / 下午) with daily status
  - Sheet 2 情况说明: employees with attendance anomalies
  - Sheet 3 月度汇总: name in column A, daily status in B–AF
  - Sheet 4 加班结算加班工资: pay section (加班费) + 调休 section (rows 90–186)
"""

from __future__ import annotations

import calendar
import logging
import os
import tempfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Sequence, Tuple, Union

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session, joinedload

from app.excel.template_generator import (
    MONTHLY_DAILY_END_COL,
    MONTHLY_DAILY_START_COL,
    MONTHLY_DATA_START_ROW,
    MONTHLY_GENERATED_ROW,
    MONTHLY_HEADER_ROW,
    MONTHLY_TITLE_ROW,
    OVERTIME_COMP_DATA_START_ROW,
    OVERTIME_COMP_DAY_START_COL,
    OVERTIME_COMP_HEADER_ROW,
    OVERTIME_COMP_MAX_ROW,
    OVERTIME_DATA_START_ROW,
    OVERTIME_DAY_COUNT,
    OVERTIME_DAY_START_COL,
    OVERTIME_MAIN_HEADER_ROW,
    OVERTIME_SETTLEMENT_COMP,
    OVERTIME_SETTLEMENT_PAY,
    SIGN_DATA_START_ROW,
    SIGN_DAY_COUNT,
    SIGN_DAY_START_COL,
    SIGN_HEADER_ROW,
    SIGN_SUMMARY_START_COL,
    SITUATION_DATA_START_ROW,
    TEMPLATE_SHEETS,
    TemplateEmployee,
    build_attendance_workbook,
    configure_monthly_summary_headers,
    format_monthly_summary_sheet,
    generate_attendance_template,
    prepare_monthly_spacer_columns,
    write_overtime_calc_formulas,
    write_overtime_compensatory_formulas,
    write_overtime_compensatory_headers,
    write_overtime_compensatory_employee_row,
    write_overtime_pay_section_headers,
    write_overtime_sheet_title,
    write_sign_sheet_legend,
    write_sign_sheet_headers,
    write_sign_sheet_summary_formulas,
)
from app.excel.field_utils import get_overtime_day_hours
from app.excel.monthly_status_display import (
    format_monthly_summary_day_status,
    monthly_summary_status_style,
)
from app.models import MonthlyAttendance
from app.services.situation_sheet import collect_situation_rows

logger = logging.getLogger(__name__)

# Recognized daily status symbols for 月度汇总 columns B–AF (day_1 … day_31).
VALID_DAY_STATUS_SYMBOLS = frozenset(
    {
        "√",
        "◇",
        "✬",
        "▼",
        "※",
        "●",
        "AL",
        "○",
        "FL",
        "ML",
        "旷工",
        "迟到",
        "缺卡",
        "正常",
        "休息",
    }
)

# Map database / DingTalk text aliases to canonical Excel display symbols.
DAY_STATUS_DISPLAY_MAP = {
    "出勤": "√",
    "present": "√",
    "正常": "正常",
    "事假": "◇",
    "personal_leave": "◇",
    "调休": "✬",
    "compensatory_leave": "✬",
    "出差": "▼",
    "business_trip": "▼",
    "病假": "※",
    "sick_leave": "※",
    "福利假": "●",
    "welfare_leave": "●",
    "年假": "AL",
    "annual_leave": "AL",
    "产假": "○",
    "陪产假": "○",
    "maternity_leave": "○",
    "丧假": "FL",
    "funeral_leave": "FL",
    "婚假": "ML",
    "marriage_leave": "ML",
    "旷工": "旷工",
    "absenteeism": "旷工",
    "旷": "旷工",
    "迟到": "迟到",
    "late": "迟到",
    "迟": "迟到",
    "缺卡": "缺卡",
    "missing_punch": "缺卡",
    "缺": "缺卡",
    "休息": "休息",
    "rest": "休息",
    "未签到": "",
}

ANOMALY_SYMBOLS = frozenset({"※", "●", "缺", "×", "迟", "旷工", "迟到", "缺卡"})
SINGLE_CHAR_DAY_STATUS_SYMBOLS = frozenset(
    symbol for symbol in VALID_DAY_STATUS_SYMBOLS if len(symbol) == 1
) | frozenset({"※", "●", "缺", "×", "迟"})
HALF_DAY_SEPARATORS = ("/", "|", "、")
MASTER_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2] / "templates" / "master_template.xlsx"
)

# Pattern rules for 签字 sheet symbols (checked in order; first match wins).
SIGN_SHEET_SYMBOL_RULES: Tuple[Tuple[Tuple[str, ...], str], ...] = (
    (("正常", "√"), "√"),
    (("旷工",), "旷工"),
    (("迟到",), "迟到"),
    (("缺卡", "未打卡"), "缺卡"),
    (("出差",), "▼"),
    (("事假",), "◇"),
    (("调休",), "✬"),
    (("病假",), "※"),
    (("年假",), "AL"),
    (("产假", "陪产假"), "○"),
    (("丧假",), "FL"),
    (("婚假",), "ML"),
    (("福利假", "生日假"), "●"),
)

SIGN_SHEET_LEGACY_SYMBOLS = frozenset(
    {"√", "◇", "✬", "▼", "※", "●", "AL", "○", "FL", "ML", "旷工", "迟到", "缺卡"}
)


@dataclass
class ExcelExportResult:
    """Generated attendance workbook on disk."""

    path: Path
    filename: str
    year: int
    month: int
    employee_count: int

    def as_stream(self) -> BytesIO:
        buffer = BytesIO(self.path.read_bytes())
        buffer.seek(0)
        return buffer

    def cleanup(self) -> None:
        try:
            self.path.unlink(missing_ok=True)
        except OSError:
            logger.warning("Failed to remove temporary Excel file: %s", self.path)


class ExcelGeneratorError(Exception):
    def __init__(self, message: str, status_code: int = 404):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


def get_day_value(record: MonthlyAttendance, day: int) -> str:
    field_name = f"day_{day}"
    overrides = record.manual_overrides or {}
    if field_name in overrides and overrides[field_name] is not None:
        return str(overrides[field_name]).strip()
    value = getattr(record, field_name, None)
    return str(value).strip() if value else ""


def _is_half_day_mark(value: str) -> bool:
    """
    Return True only when *value* encodes two distinct half-day marks.

    Full-day statuses such as ``旷工`` / ``迟到`` / ``缺卡`` must not be split.
    """
    text = value.strip()
    if not text:
        return False
    if any(separator in text for separator in HALF_DAY_SEPARATORS):
        return True
    if text in VALID_DAY_STATUS_SYMBOLS or text in DAY_STATUS_DISPLAY_MAP:
        return False
    return (
        len(text) == 2
        and text[0] in SINGLE_CHAR_DAY_STATUS_SYMBOLS
        and text[1] in SINGLE_CHAR_DAY_STATUS_SYMBOLS
    )


def split_day_halves(value: str) -> Tuple[str, str]:
    if not value:
        return "", ""
    text = value.strip()
    if not _is_half_day_mark(text):
        return text, text
    for separator in HALF_DAY_SEPARATORS:
        if separator in text:
            left, right = text.split(separator, 1)
            return left.strip(), right.strip()
    if (
        len(text) == 2
        and text[0] in SINGLE_CHAR_DAY_STATUS_SYMBOLS
        and text[1] in SINGLE_CHAR_DAY_STATUS_SYMBOLS
    ):
        return text[0], text[1]
    return text, text


def combined_day_status(am_value: str, pm_value: str) -> str:
    if am_value and pm_value and am_value == pm_value:
        return am_value
    if am_value and pm_value:
        return f"{am_value}{pm_value}"
    return am_value or pm_value


def _has_anomaly_in_day(value: str) -> bool:
    if not value:
        return False
    am_value, pm_value = split_day_halves(value)
    return am_value in ANOMALY_SYMBOLS or pm_value in ANOMALY_SYMBOLS


def first_anomaly_date(record: MonthlyAttendance, year: int, month: int) -> Optional[str]:
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        if _has_anomaly_in_day(get_day_value(record, day)):
            return f"{year}-{month:02d}-{day:02d}"
    return None


def resolve_day_value(record: MonthlyAttendance, day: int) -> str:
    return get_day_value(record, day) or ""


def map_day_status_to_excel(raw: str) -> str:
    """Normalize a single half-day or full-day status string for Excel display."""
    if not raw:
        return ""
    text = raw.strip()
    if not text:
        return ""
    if text in VALID_DAY_STATUS_SYMBOLS:
        return text
    mapped = DAY_STATUS_DISPLAY_MAP.get(text)
    if mapped is not None:
        return mapped
    return text


def map_sign_sheet_status(raw: str, rules=None) -> str:
    """
    Map a daily status value to a 签字 sheet symbol (√, ◇, ▼, …).

    Uses configurable rules when provided; otherwise falls back to built-in defaults.
    """
    from app.services.attendance_rule_engine import default_rules, map_status_to_symbol

    resolved = rules if rules is not None else default_rules()
    return map_status_to_symbol(raw, resolved)


def _prepare_sign_employee_rows(ws, am_row: int, pm_row: int, name: str) -> None:
    """Unmerge template merges and set blank 签名 + duplicate 姓名 on both rows."""
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row <= pm_row
            and merged.max_row >= am_row
            and merged.min_col <= 2
            and merged.max_col >= 1
        ):
            ws.unmerge_cells(str(merged))

    ws.cell(row=am_row, column=1, value=None)
    ws.cell(row=pm_row, column=1, value=None)
    ws.cell(row=am_row, column=2, value=name)
    ws.cell(row=pm_row, column=2, value=name)
    ws.cell(row=am_row, column=3, value="上午")
    ws.cell(row=pm_row, column=3, value="下午")


def format_export_day_status(
    record: MonthlyAttendance,
    year: int,
    month: int,
    day: int,
) -> Optional[str]:
    """
    Read ``day_{day}`` from *record* and return the Excel display symbol.

    Empty / null database values become blank cells (``None``).
    """
    days_in_month = calendar.monthrange(year, month)[1]
    if day > days_in_month:
        return None

    day_value = resolve_day_value(record, day)
    if not day_value:
        return None

    text = day_value.strip()
    if not _is_half_day_mark(text):
        display = map_day_status_to_excel(text)
        return display if display else None

    am_value, pm_value = split_day_halves(text)
    mapped_am = map_day_status_to_excel(am_value)
    mapped_pm = map_day_status_to_excel(pm_value)
    display = combined_day_status(mapped_am, mapped_pm)
    return display if display else None


def populate_employee_daily_columns(
    ws,
    row: int,
    record: MonthlyAttendance,
    year: int,
    month: int,
) -> None:
    """
    Write ``day_1`` … ``day_31`` from *record* to columns B–AF on *row*.

    Column index: B = day 1, C = day 2, …, AF = day 31.
    Uses DingTalk descriptive text with per-status background colors.
    """
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, SIGN_DAY_COUNT + 1):
        col = MONTHLY_DAILY_START_COL + day - 1
        cell = ws.cell(row=row, column=col)
        if day > days_in_month:
            cell.value = None
            continue
        display_value = format_monthly_summary_day_status(
            record,
            year,
            month,
            day,
            resolve_day_value=resolve_day_value,
        )
        cell.value = display_value
        if display_value:
            fill, font = monthly_summary_status_style(display_value)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font


def populate_monthly_daily_status_from_db(
    ws,
    records: Sequence[MonthlyAttendance],
    year: int,
    month: int,
    *,
    start_row: int = MONTHLY_DATA_START_ROW,
) -> None:
    """
    Populate daily attendance for all employees on the 月度汇总 sheet.

    Queries are performed upstream; *records* should be ``monthly_attendance`` rows
    (with ``employee`` loaded) for the target year/month. Data is written starting
    at *start_row* (default row 5), columns B–AF.
    """
    for index, record in enumerate(records):
        row = start_row + index
        populate_employee_daily_columns(ws, row, record, year, month)


def query_attendance_records(
    db: Session,
    company_id: int,
    year: int,
    month: int,
) -> List[MonthlyAttendance]:
    return (
        db.query(MonthlyAttendance)
        .options(joinedload(MonthlyAttendance.employee))
        .filter(
            MonthlyAttendance.company_id == company_id,
            MonthlyAttendance.year == year,
            MonthlyAttendance.month == month,
        )
        .order_by(MonthlyAttendance.employee_id)
        .all()
    )


def _record_to_template_employee(record: MonthlyAttendance) -> TemplateEmployee:
    employee = record.employee
    return TemplateEmployee(
        name=employee.name,
        department=employee.department or "",
        position=employee.position or "",
        employee_code=employee.employee_code or "",
    )


def _count_template_employee_slots(sign_ws) -> int:
    count = 0
    row = SIGN_DATA_START_ROW
    while True:
        if sign_ws.cell(row=row, column=3).value != "上午":
            break
        if sign_ws.cell(row=row + 1, column=3).value != "下午":
            break
        count += 1
        row += 2
    return count


def load_workbook_from_template(
    year: int,
    month: int,
    employees: Sequence[TemplateEmployee],
    template_path: Optional[Path] = None,
):
    path = template_path or MASTER_TEMPLATE_PATH
    employee_count = len(employees)

    if path.exists():
        workbook = load_workbook(path)
        if workbook.sheetnames == list(TEMPLATE_SHEETS):
            sign_ws = workbook[TEMPLATE_SHEETS[0]]
            monthly_ws = workbook[TEMPLATE_SHEETS[2]]
            slot_count = _count_template_employee_slots(sign_ws)
            monthly_cols = monthly_ws.max_column or 0
            if (
                slot_count >= employee_count
                and monthly_cols >= MONTHLY_DAILY_END_COL
            ):
                logger.info(
                    "Loaded master template from %s (%s slots, %s employees)",
                    path,
                    slot_count,
                    employee_count,
                )
                return workbook
        logger.info(
            "Rebuilding workbook for %s employees (template mismatch or insufficient slots)",
            employee_count,
        )
    else:
        logger.warning("Master template not found at %s; building workbook", path)

    return build_attendance_workbook(year, month, employees)


def populate_sign_sheet(
    ws,
    records: Sequence[MonthlyAttendance],
    year: int,
    month: int,
    *,
    rules=None,
) -> None:
    """Sheet 1 签字: two rows per employee (上午 / 下午) with identical symbols in D–AH."""
    write_sign_sheet_headers(ws, year, month)
    write_sign_sheet_legend(ws)
    days_in_month = calendar.monthrange(year, month)[1]
    for index, record in enumerate(records):
        am_row = SIGN_DATA_START_ROW + index * 2
        pm_row = am_row + 1
        name = record.employee.name
        _prepare_sign_employee_rows(ws, am_row, pm_row, name)

        for day in range(1, SIGN_DAY_COUNT + 1):
            col = SIGN_DAY_START_COL + day - 1
            if day > days_in_month:
                ws.cell(row=am_row, column=col, value=None)
                ws.cell(row=pm_row, column=col, value=None)
                continue
            day_value = resolve_day_value(record, day)
            symbol = map_sign_sheet_status(day_value, rules)
            if not symbol:
                ws.cell(row=am_row, column=col, value=None)
                ws.cell(row=pm_row, column=col, value=None)
                continue
            ws.cell(row=am_row, column=col, value=symbol)
            ws.cell(row=pm_row, column=col, value=symbol)

    write_sign_sheet_summary_formulas(ws, len(records), year, month)


def populate_monthly_sheet(
    ws,
    records: Sequence[MonthlyAttendance],
    year: int,
    month: int,
    *,
    generated_at: Optional[datetime] = None,
) -> None:
    """Sheet 3 月度汇总: name in column A, daily status in B–AF."""
    configure_monthly_summary_headers(ws, year, month, generated_at=generated_at)

    for index, record in enumerate(records):
        row = MONTHLY_DATA_START_ROW + index
        employee = record.employee
        ws.cell(row=row, column=1, value=employee.name)

    populate_monthly_daily_status_from_db(ws, records, year, month)

    last_row = MONTHLY_DATA_START_ROW + len(records) - 1
    format_monthly_summary_sheet(ws, year, month, last_row=last_row)
    prepare_monthly_spacer_columns(ws, last_row=last_row)
    ws.freeze_panes = ws.cell(row=MONTHLY_DATA_START_ROW, column=MONTHLY_DAILY_START_COL).coordinate


def populate_situation_sheet(ws, records: Sequence[MonthlyAttendance], year: int, month: int) -> None:
    """Sheet 2 情况说明: one row per employee — same data as the web UI explanation tab."""
    situation_rows = collect_situation_rows(
        records,
        year,
        month,
        first_anomaly_date=first_anomaly_date,
    )

    row = SITUATION_DATA_START_ROW
    for item in situation_rows:
        ws.cell(row=row, column=1, value=item.name)
        ws.cell(row=row, column=2, value=item.date)
        ws.cell(row=row, column=3, value=item.anomaly)
        ws.cell(row=row, column=4, value=item.supplement_submitted or None)
        ws.cell(row=row, column=5, value=item.notes or None)
        row += 1


def _write_overtime_daily_cells(
    ws,
    row: int,
    record: MonthlyAttendance,
    year: int,
    month: int,
    *,
    start_col: int,
) -> None:
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, OVERTIME_DAY_COUNT + 1):
        col = start_col + day - 1
        if day > days_in_month:
            ws.cell(row=row, column=col, value=None)
            continue
        hours = get_overtime_day_hours(record, day)
        ws.cell(row=row, column=col, value=hours if hours else None)


def populate_overtime_sheet(
    ws,
    records: Sequence[MonthlyAttendance],
    year: int,
    month: int,
) -> None:
    write_overtime_sheet_title(ws, year, month)
    write_overtime_pay_section_headers(ws, year, month)
    for index, record in enumerate(records):
        row = OVERTIME_DATA_START_ROW + index
        employee = record.employee
        ws.cell(row=row, column=1, value=employee.name)
        ws.cell(row=row, column=2, value=employee.department or "")
        ws.cell(row=row, column=3, value=OVERTIME_SETTLEMENT_PAY)
        _write_overtime_daily_cells(
            ws,
            row,
            record,
            year,
            month,
            start_col=OVERTIME_DAY_START_COL,
        )

    write_overtime_calc_formulas(ws, len(records))

    write_overtime_compensatory_headers(ws, year, month)
    comp_slots = OVERTIME_COMP_MAX_ROW - OVERTIME_COMP_DATA_START_ROW + 1
    for index in range(comp_slots):
        row = OVERTIME_COMP_DATA_START_ROW + index
        if index < len(records):
            record = records[index]
            employee = record.employee
            write_overtime_compensatory_employee_row(
                ws,
                row,
                name=employee.name,
                department=employee.department or "",
            )
            _write_overtime_daily_cells(
                ws,
                row,
                record,
                year,
                month,
                start_col=OVERTIME_COMP_DAY_START_COL,
            )
        else:
            ws.cell(row=row, column=3, value=OVERTIME_SETTLEMENT_COMP)
            write_overtime_compensatory_formulas(ws, row)


def populate_workbook(
    workbook,
    records: Sequence[MonthlyAttendance],
    year: int,
    month: int,
    *,
    generated_at: Optional[datetime] = None,
    rules=None,
) -> None:
    sign_ws = workbook[TEMPLATE_SHEETS[0]]
    situation_ws = workbook[TEMPLATE_SHEETS[1]]
    monthly_ws = workbook[TEMPLATE_SHEETS[2]]
    if monthly_ws.title != TEMPLATE_SHEETS[2]:
        monthly_ws.title = TEMPLATE_SHEETS[2]
    overtime_ws = workbook[TEMPLATE_SHEETS[3]]

    populate_sign_sheet(sign_ws, records, year, month, rules=rules)
    populate_situation_sheet(situation_ws, records, year, month)
    populate_monthly_sheet(monthly_ws, records, year, month, generated_at=generated_at)
    populate_overtime_sheet(overtime_ws, records, year, month)


def generate_attendance_excel(
    db: Session,
    company_id: int,
    year: int,
    month: int,
    *,
    template_path: Optional[Union[str, Path]] = None,
    output_dir: Optional[Union[str, Path]] = None,
) -> ExcelExportResult:
    """
    Load the master template, populate sheets from monthly_attendance, and save to disk.
    """
    if month < 1 or month > 12:
        raise ExcelGeneratorError("Month must be between 1 and 12", status_code=400)

    records = query_attendance_records(db, company_id, year, month)
    if not records:
        raise ExcelGeneratorError(f"No attendance data for {year}-{month:02d}", status_code=404)

    employees = [_record_to_template_employee(record) for record in records]
    resolved_template = Path(template_path) if template_path else None

    workbook = load_workbook_from_template(year, month, employees, resolved_template)
    generated_at = datetime.utcnow()
    populate_workbook(workbook, records, year, month, generated_at=generated_at)

    filename = f"attendance_{year}_{month:02d}.xlsx"
    if output_dir:
        target_dir = Path(output_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        output_path = target_dir / filename
        workbook.save(output_path)
    else:
        fd, temp_name = tempfile.mkstemp(prefix="attendance_", suffix=".xlsx")
        os.close(fd)
        output_path = Path(temp_name)
        workbook.save(output_path)

    logger.info(
        "Generated attendance Excel for company_id=%s %s-%02d (%s employees) -> %s",
        company_id,
        year,
        month,
        len(records),
        output_path,
    )

    return ExcelExportResult(
        path=output_path,
        filename=filename,
        year=year,
        month=month,
        employee_count=len(records),
    )


def generate_monthly_summary_excel(
    db: Session,
    company_id: int,
    year: int,
    month: int,
    *,
    output_dir: Optional[Union[str, Path]] = None,
) -> ExcelExportResult:
    """
    Generate a single-sheet workbook that contains only 月度汇总.
    """
    if month < 1 or month > 12:
        raise ExcelGeneratorError("Month must be between 1 and 12", status_code=400)

    records = query_attendance_records(db, company_id, year, month)
    if not records:
        raise ExcelGeneratorError(f"No attendance data for {year}-{month:02d}", status_code=404)

    workbook = Workbook()
    ws = workbook.active
    ws.title = TEMPLATE_SHEETS[2]
    populate_monthly_sheet(ws, records, year, month, generated_at=datetime.utcnow())

    filename = f"monthly_summary_{year}_{month:02d}.xlsx"
    if output_dir:
        target_dir = Path(output_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        output_path = target_dir / filename
        workbook.save(output_path)
    else:
        fd, temp_name = tempfile.mkstemp(prefix="monthly_summary_", suffix=".xlsx")
        os.close(fd)
        output_path = Path(temp_name)
        workbook.save(output_path)

    logger.info(
        "Generated monthly summary Excel for company_id=%s %s-%02d (%s employees) -> %s",
        company_id,
        year,
        month,
        len(records),
        output_path,
    )

    return ExcelExportResult(
        path=output_path,
        filename=filename,
        year=year,
        month=month,
        employee_count=len(records),
    )


def ensure_master_template(path: Optional[Union[str, Path]] = None) -> Path:
    """Create or refresh the master template when missing or outdated."""
    target = Path(path) if path else MASTER_TEMPLATE_PATH
    needs_refresh = not target.exists()
    if not needs_refresh:
        from openpyxl import load_workbook as _load_wb

        wb = _load_wb(target, read_only=True)
        try:
            if TEMPLATE_SHEETS[2] in wb.sheetnames:
                monthly_ws = wb[TEMPLATE_SHEETS[2]]
                monthly_cols = monthly_ws.max_column or 0
                title = monthly_ws.cell(row=MONTHLY_TITLE_ROW, column=1).value
                timestamp = monthly_ws.cell(row=MONTHLY_GENERATED_ROW, column=1).value
                header_col2 = monthly_ws.cell(row=MONTHLY_HEADER_ROW, column=2).value
                sign_header = None
                if TEMPLATE_SHEETS[0] in wb.sheetnames:
                    sign_ws = wb[TEMPLATE_SHEETS[0]]
                    sign_header = sign_ws.cell(
                        row=SIGN_HEADER_ROW,
                        column=SIGN_SUMMARY_START_COL,
                    ).value
                needs_refresh = (
                    monthly_cols < MONTHLY_DAILY_END_COL
                    or header_col2 in ("考勤组", "部门", "工号", "职位")
                    or not (isinstance(title, str) and title.startswith("月度汇总"))
                    or not (
                        isinstance(timestamp, str) and timestamp.startswith("报表生成时间：")
                    )
                    or sign_header != "出勤合计"
                )
            if not needs_refresh and TEMPLATE_SHEETS[3] in wb.sheetnames:
                overtime_ws = wb[TEMPLATE_SHEETS[3]]
                comp_header = overtime_ws.cell(row=OVERTIME_COMP_HEADER_ROW, column=1).value
                pay_summary = overtime_ws.cell(row=OVERTIME_MAIN_HEADER_ROW, column=36).value
                pay_header_fill = overtime_ws.cell(row=OVERTIME_MAIN_HEADER_ROW, column=36).fill.fgColor.rgb
                needs_refresh = (
                    comp_header != "姓名"
                    or not (
                        isinstance(pay_summary, str) and "加班时长合计" in pay_summary
                    )
                    or pay_header_fill == "0090EE90"
                )
        finally:
            wb.close()
    if needs_refresh:
        generate_attendance_template(target, year=datetime.utcnow().year, month=datetime.utcnow().month)
    return target
